from django.shortcuts import render, redirect, get_object_or_404
from django.http import HttpResponse, JsonResponse
from django.contrib.auth.decorators import login_required
from django.views.decorators.csrf import csrf_exempt
from .finance_manager import finance_dashboard, finance_projects, finance_cost_management, finance_notifications
from .engineer_management import (
    engineer_list, engineer_create, engineer_detail,
    engineer_edit, engineer_deactivate, engineer_activate,
    engineer_delete
)
from .budget_notifications import forward_budget_alert_to_finance_view
from projeng.models import Project, ProjectProgress, ProjectCost
from django.contrib.auth.models import Group
from django.core.paginator import Paginator
from django.template.loader import get_template
from django.utils import timezone
import os
import json
import csv
import io
from datetime import datetime
import openpyxl
# Optional PDF libraries
try:
    from xhtml2pdf import pisa
except Exception:
    pisa = None
try:
    import pdfkit  # requires wkhtmltopdf binary installed
except Exception:
    pdfkit = None
from django.conf import settings
from collections import Counter, defaultdict
from monitoring.forms import ProjectForm

# Import centralized access control functions (MUST be before views that use them)
from gistagum.access_control import (
    is_head_engineer, 
    is_finance_manager, 
    is_project_engineer,
    is_project_or_head_engineer,
    is_finance_or_head_engineer,
    head_engineer_required,
    prevent_project_engineer_access,
    get_user_dashboard_url
)

def _get_or_create_project_type_by_name(name: str):
    """
    Create (or return existing) ProjectType by name.
    Used by the "Add new project type" UX in modals.
    """
    from django.utils.text import slugify
    from projeng.models import ProjectType

    cleaned = (name or "").strip()
    if not cleaned:
        return None

    existing = ProjectType.objects.filter(name__iexact=cleaned).first()
    if existing:
        return existing

    # Generate a stable unique code
    base = (slugify(cleaned) or "custom").replace("-", "_")
    base = (base or "custom")[:50]
    code = base
    if ProjectType.objects.filter(code__iexact=code).exists():
        i = 2
        while True:
            suffix = f"_{i}"
            code = (base[: max(1, 50 - len(suffix))] + suffix)[:50]
            if not ProjectType.objects.filter(code__iexact=code).exists():
                break
            i += 1

    # Use sensible defaults for required characteristics
    return ProjectType.objects.create(
        name=cleaned[:100],
        code=code,
        description="",
        density_level="medium",
        height_category=None,
        requires_industrial=False,
        requires_commercial=False,
        requires_residential=False,
    )

def home(request):
    """
    Redirect to login page - no direct access to home
    """
    from django.shortcuts import redirect
    return redirect('login')

@login_required
@prevent_project_engineer_access
def dashboard(request):
    try:
        print(f"Dashboard view accessed by user: {request.user.username}, authenticated: {request.user.is_authenticated}")
        from projeng.models import Project, ProjectProgress
        from django.utils import timezone
        from django.db.models import Max
        from collections import Counter, defaultdict
        
        # Role-based queryset
        if is_head_engineer(request.user) or is_finance_manager(request.user):
            print(f"User {request.user.username} is head engineer or finance manager, showing all projects")
            projects = Project.objects.all()
        elif is_project_engineer(request.user):
            projects = Project.objects.filter(assigned_engineers=request.user)
        else:
            projects = Project.objects.none()
        # Recent projects (5 most recent)
        recent_projects = projects.order_by('-created_at')[:5]
        
        # Calculate metrics with dynamic delayed status
        today = timezone.now().date()
        project_count = projects.count()
        
        # Get latest progress for all projects
        project_ids = [p.id for p in projects]
        latest_progress = {}
        if project_ids:
            latest_progress_qs = ProjectProgress.objects.filter(
                project_id__in=project_ids
            ).values('project_id').annotate(
                latest_date=Max('date'),
                latest_created=Max('created_at')
            )
            for item in latest_progress_qs:
                latest = ProjectProgress.objects.filter(
                    project_id=item['project_id'],
                    date=item['latest_date'],
                    created_at=item['latest_created']
                ).order_by('-created_at').first()
                if latest and latest.percentage_complete is not None:
                    latest_progress[item['project_id']] = int(latest.percentage_complete)
        
        # Calculate status counts dynamically
        completed_count = 0
        in_progress_count = 0
        planned_count = 0
        delayed_count = 0
        
        for p in projects:
            progress = latest_progress.get(p.id, 0)
            stored_status = p.status or ''
            
            # Calculate actual status dynamically
            # Priority: completed > delayed > in_progress > planned
            if progress >= 99:
                completed_count += 1
            elif stored_status == 'delayed':
                # Already marked as delayed in database
                delayed_count += 1
            elif progress < 99 and p.end_date and p.end_date < today and stored_status in ['in_progress', 'ongoing']:
                # Project is delayed if: end_date passed, progress < 99%, and status is in_progress/ongoing
                delayed_count += 1
            elif stored_status in ['in_progress', 'ongoing']:
                in_progress_count += 1
            elif stored_status in ['planned', 'pending']:
                planned_count += 1
            elif stored_status == 'completed':
                completed_count += 1
        # Projects created per month (last 12 months)
        from django.db.models import Count
        from django.db.models.functions import TruncMonth
        from datetime import timedelta

        twelve_months_ago = timezone.now() - timedelta(days=365)
        created_by_month_qs = (
            projects.filter(created_at__gte=twelve_months_ago)
            .annotate(month=TruncMonth('created_at'))
            .values('month')
            .annotate(total=Count('id'))
            .order_by('month')
        )
        projects_created_labels = [row['month'].strftime('%b %Y') for row in created_by_month_qs if row.get('month')]
        projects_created_counts = [int(row['total']) for row in created_by_month_qs]
        if not projects_created_labels:
            projects_created_labels = ['No Data']
            projects_created_counts = [0]

        # Collaborative analytics
        collab_by_barangay = defaultdict(int)
        collab_by_status = defaultdict(int)
        for p in projects:
            if p.barangay and isinstance(p.barangay, str) and p.barangay.strip():
                collab_by_barangay[p.barangay.strip()] += 1
            if p.status and isinstance(p.status, str) and p.status.strip():
                # Use display-friendly status
                status = (
                    'Ongoing' if p.status in ['in_progress', 'ongoing'] else
                    'Planned' if p.status in ['planned', 'pending'] else
                    'Completed' if p.status == 'completed' else
                    'Delayed' if p.status == 'delayed' else p.status.title()
                )
                collab_by_status[status] += 1
        # Sort keys for consistent chart order
        collab_by_barangay = {k: collab_by_barangay[k] for k in sorted(collab_by_barangay.keys())}
        collab_by_status = {k: collab_by_status[k] for k in sorted(collab_by_status.keys())}

        # Calculate completion rate
        completion_rate = 0.0
        if project_count > 0:
            completion_rate = (completed_count / project_count) * 100
        
        context = {
            'recent_projects': recent_projects,
            'project_count': project_count,
            'completed_count': completed_count,
            'in_progress_count': in_progress_count,
            'planned_count': planned_count,
            'delayed_count': delayed_count,
            'completion_rate': round(completion_rate, 1),
            'collab_by_barangay': dict(collab_by_barangay),
            'collab_by_status': dict(collab_by_status),
            'projects_created_labels': projects_created_labels,
            'projects_created_counts': projects_created_counts,
        }
        return render(request, 'monitoring/dashboard.html', context)
    except Exception as e:
        import logging
        logger = logging.getLogger(__name__)
        logger.error(f'Error in dashboard view: {str(e)}', exc_info=True)
        from django.http import HttpResponseServerError
        return HttpResponseServerError(f'Server Error: {str(e)}')

@login_required
@prevent_project_engineer_access
def dashboard_budget_utilization_data(request):
    """API endpoint for Budget Utilization by Project chart (supports period: week/month/year)"""
    from projeng.models import Project, ProjectCost
    from django.db.models import Sum
    from django.http import JsonResponse
    from datetime import timedelta

    if is_head_engineer(request.user) or is_finance_manager(request.user):
        projects = Project.objects.filter(project_cost__isnull=False).exclude(project_cost=0)
    else:
        projects = Project.objects.none()

    # Optional period filter: only include costs within the selected period
    period = (request.GET.get('period') or 'month').strip().lower()
    now = timezone.now().date()
    if period == 'week':
        start_date = now - timedelta(days=7)
    elif period == 'year':
        start_date = now - timedelta(days=365)
    else:
        start_date = now - timedelta(days=30)

    # Calculate budget utilization for each project (costs within period vs full budget)
    project_data = []
    for project in projects:
        total_cost = ProjectCost.objects.filter(
            project=project,
            date__gte=start_date
        ).aggregate(total=Sum('amount'))['total'] or 0

        if project.project_cost and float(project.project_cost) > 0:
            utilization = (float(total_cost) / float(project.project_cost)) * 100
            project_data.append({
                'name': project.name[:25] + '...' if len(project.name) > 25 else project.name,
                'utilization': round(utilization, 1),
                'total_cost': float(total_cost),
                'budget': float(project.project_cost)
            })
    
    # Sort by utilization (highest first) and take top 10
    project_data.sort(key=lambda x: x['utilization'], reverse=True)
    project_data = project_data[:10]
    
    labels = [p['name'] for p in project_data]
    data = [p['utilization'] for p in project_data]
    
    # Color coding: Red (>100%), Orange (80-100%), Green (<80%)
    background_colors = [
        'rgba(239, 68, 68, 0.7)' if d > 100 else
        'rgba(251, 146, 60, 0.7)' if d > 80 else
        'rgba(34, 197, 94, 0.7)'
        for d in data
    ]
    border_colors = [
        'rgba(239, 68, 68, 1)' if d > 100 else
        'rgba(251, 146, 60, 1)' if d > 80 else
        'rgba(34, 197, 94, 1)'
        for d in data
    ]
    
    return JsonResponse({
        'labels': labels,
        'datasets': [{
            'label': 'Budget Utilization (%)',
            'data': data,
            'backgroundColor': background_colors,
            'borderColor': border_colors,
            'borderWidth': 1
        }]
    })


@login_required
@prevent_project_engineer_access
def dashboard_collab_analytics_data(request):
    """API endpoint for Projects per Barangay and Projects per Status, filtered by period (week/month/year)."""
    from projeng.models import Project
    from django.http import JsonResponse
    from datetime import timedelta
    from collections import defaultdict

    if not (is_head_engineer(request.user) or is_finance_manager(request.user)):
        return JsonResponse({'barangay': {}, 'status': {}})

    period = (request.GET.get('period') or 'month').strip().lower()
    now = timezone.now().date()
    if period == 'week':
        start_date = now - timedelta(days=7)
    elif period == 'year':
        start_date = now - timedelta(days=365)
    else:
        start_date = now - timedelta(days=30)

    projects = Project.objects.filter(created_at__date__gte=start_date)
    collab_by_barangay = defaultdict(int)
    collab_by_status = defaultdict(int)
    for p in projects:
        if p.barangay and isinstance(p.barangay, str) and p.barangay.strip():
            collab_by_barangay[p.barangay.strip()] += 1
        if p.status and isinstance(p.status, str) and p.status.strip():
            status = (
                'Ongoing' if p.status in ['in_progress', 'ongoing'] else
                'Planned' if p.status in ['planned', 'pending'] else
                'Completed' if p.status == 'completed' else
                'Delayed' if p.status == 'delayed' else p.status.title()
            )
            collab_by_status[status] += 1
    collab_by_barangay = {k: collab_by_barangay[k] for k in sorted(collab_by_barangay.keys())}
    collab_by_status = {k: collab_by_status[k] for k in sorted(collab_by_status.keys())}
    return JsonResponse({'barangay': collab_by_barangay, 'status': dict(collab_by_status)})


@login_required
@prevent_project_engineer_access
def dashboard_cost_breakdown_data(request):
    """API endpoint for Cost Breakdown by Type chart"""
    from projeng.models import ProjectCost
    from django.db.models import Sum
    from django.http import JsonResponse
    from datetime import timedelta
    
    if is_head_engineer(request.user) or is_finance_manager(request.user):
        period = (request.GET.get('period') or 'month').strip().lower()
        now = timezone.now().date()
        if period == 'week':
            start_date = now - timedelta(days=7)
        elif period == 'year':
            start_date = now - timedelta(days=365)
        else:
            start_date = now - timedelta(days=30)

        # Get cost breakdown by type within selected period
        cost_by_type = ProjectCost.objects.filter(date__gte=start_date).values('cost_type').annotate(
            total=Sum('amount')
        ).order_by('-total')
    else:
        cost_by_type = []
    
    labels = [item['cost_type'].title() for item in cost_by_type]
    data = [float(item['total']) for item in cost_by_type]
    
    # Color scheme for cost types
    colors = {
        'Material': 'rgba(59, 130, 246, 0.7)',    # Blue
        'Labor': 'rgba(251, 191, 36, 0.7)',       # Yellow
        'Equipment': 'rgba(34, 197, 94, 0.7)',    # Green
        'Other': 'rgba(139, 92, 246, 0.7)'        # Purple
    }
    
    background_colors = [colors.get(label, 'rgba(156, 163, 175, 0.7)') for label in labels]
    border_colors = [bg.replace('0.7', '1') for bg in background_colors]
    
    return JsonResponse({
        'labels': labels,
        'datasets': [{
            'label': 'Total Cost (₱)',
            'data': data,
            'backgroundColor': background_colors,
            'borderColor': border_colors,
            'borderWidth': 2
        }]
    })

@login_required
@prevent_project_engineer_access
def dashboard_monthly_spending_data(request):
    """API endpoint for Spending Trend chart (week/month/year)"""
    from projeng.models import ProjectCost
    from django.db.models import Sum
    from django.db.models.functions import TruncMonth, TruncYear
    from django.http import JsonResponse
    from datetime import timedelta
    from django.utils import timezone
    
    if is_head_engineer(request.user) or is_finance_manager(request.user):
        period = (request.GET.get('period') or 'month').strip().lower()

        if period == 'year':
            start = timezone.now() - timedelta(days=365 * 5)
            qs = (
                ProjectCost.objects.filter(date__gte=start)
                .annotate(bucket=TruncYear('date'))
                .values('bucket')
                .annotate(total=Sum('amount'))
                .order_by('bucket')
            )
            labels = [item['bucket'].strftime('%Y') for item in qs if item.get('bucket')]
            totals = [float(item['total']) for item in qs]
            series_label = 'Yearly Spending (₱)'
        elif period == 'week':
            # TruncWeek isn't available in all Django versions; fall back to day if missing
            try:
                from django.db.models.functions import TruncWeek  # type: ignore
                start = timezone.now() - timedelta(days=7 * 12)
                qs = (
                    ProjectCost.objects.filter(date__gte=start)
                    .annotate(bucket=TruncWeek('date'))
                    .values('bucket')
                    .annotate(total=Sum('amount'))
                    .order_by('bucket')
                )
                labels = [item['bucket'].strftime('Wk of %b %d') for item in qs if item.get('bucket')]
                totals = [float(item['total']) for item in qs]
                series_label = 'Weekly Spending (₱)'
            except Exception:
                start = timezone.now() - timedelta(days=7 * 12)
                qs = (
                    ProjectCost.objects.filter(date__gte=start)
                    .values('date')
                    .annotate(total=Sum('amount'))
                    .order_by('date')
                )
                labels = [str(item['date']) for item in qs if item.get('date')]
                totals = [float(item['total']) for item in qs]
                series_label = 'Daily Spending (₱)'
        else:
            # Default: last 12 months
            start = timezone.now() - timedelta(days=365)
            qs = (
                ProjectCost.objects.filter(date__gte=start)
                .annotate(bucket=TruncMonth('date'))
                .values('bucket')
                .annotate(total=Sum('amount'))
                .order_by('bucket')
            )
            labels = [item['bucket'].strftime('%b %Y') for item in qs if item.get('bucket')]
            totals = [float(item['total']) for item in qs]
            series_label = 'Monthly Spending (₱)'
    else:
        labels, totals, series_label = [], [], 'Spending (₱)'
    
    if not labels:
        labels = ['No Data']
        totals = [0]
    
    return JsonResponse({
        'labels': labels,
        'datasets': [{
            'label': series_label,
            'data': totals,
            'borderColor': 'rgba(59, 130, 246, 1)',
            'backgroundColor': 'rgba(59, 130, 246, 0.1)',
            'tension': 0.4,
            'fill': True,
            'borderWidth': 2
        }]
    })


@login_required
@prevent_project_engineer_access
def dashboard_projects_created_data(request):
    """API endpoint for Projects Created per Month chart (week/month/year)."""
    from projeng.models import Project
    from django.db.models import Count
    from django.db.models.functions import TruncMonth, TruncWeek, TruncYear
    from django.http import JsonResponse
    from datetime import timedelta

    if not (is_head_engineer(request.user) or is_finance_manager(request.user)):
        return JsonResponse({'labels': ['No Data'], 'datasets': [{'label': 'Projects Created', 'data': [0]}]})

    period = (request.GET.get('period') or 'month').strip().lower()
    now = timezone.now()
    if period == 'week':
        start = now - timedelta(days=7 * 12)
        try:
            qs = (
                Project.objects.filter(created_at__gte=start)
                .annotate(bucket=TruncWeek('created_at'))
                .values('bucket')
                .annotate(total=Count('id'))
                .order_by('bucket')
            )
            labels = [item['bucket'].strftime('Wk of %b %d') for item in qs if item.get('bucket')]
            counts = [int(item['total']) for item in qs]
        except Exception:
            qs = (
                Project.objects.filter(created_at__gte=start)
                .annotate(bucket=TruncMonth('created_at'))
                .values('bucket')
                .annotate(total=Count('id'))
                .order_by('bucket')
            )
            labels = [item['bucket'].strftime('%b %Y') for item in qs if item.get('bucket')]
            counts = [int(item['total']) for item in qs]
    elif period == 'year':
        start = now - timedelta(days=365 * 5)
        qs = (
            Project.objects.filter(created_at__gte=start)
            .annotate(bucket=TruncYear('created_at'))
            .values('bucket')
            .annotate(total=Count('id'))
            .order_by('bucket')
        )
        labels = [item['bucket'].strftime('%Y') for item in qs if item.get('bucket')]
        counts = [int(item['total']) for item in qs]
    else:
        start = now - timedelta(days=365)
        qs = (
            Project.objects.filter(created_at__gte=start)
            .annotate(bucket=TruncMonth('created_at'))
            .values('bucket')
            .annotate(total=Count('id'))
            .order_by('bucket')
        )
        labels = [item['bucket'].strftime('%b %Y') for item in qs if item.get('bucket')]
        counts = [int(item['total']) for item in qs]

    if not labels:
        labels = ['No Data']
        counts = [0]

    return JsonResponse({
        'labels': labels,
        'datasets': [{
            'label': 'Projects Created',
            'data': counts,
            'backgroundColor': 'rgba(100, 116, 139, 0.7)',
            'borderColor': 'rgba(100, 116, 139, 1)',
            'borderWidth': 1
        }]
    })


@login_required
@head_engineer_required
def project_list(request):
    from projeng.models import SourceOfFunds, ProjectType
    
    if request.method == 'POST':
        # Check if this is an AJAX request
        is_ajax = request.headers.get('X-Requested-With') == 'XMLHttpRequest' or request.content_type == 'application/json'
        
        # Allow "Add new source of funds" while keeping a dropdown UI
        post_data = request.POST.copy()
        new_source_of_funds = (post_data.get('source_of_funds_new') or '').strip()
        selected_source_of_funds = (post_data.get('source_of_funds') or '').strip()
        # Treat __add_new__ or empty as "use source_of_funds_new" when user typed a new value
        if new_source_of_funds and (not selected_source_of_funds or selected_source_of_funds == '__add_new__'):
            # Case-insensitive de-dupe
            existing = SourceOfFunds.objects.filter(name__iexact=new_source_of_funds).first()
            if not existing:
                SourceOfFunds.objects.create(name=new_source_of_funds)
            post_data['source_of_funds'] = new_source_of_funds
        elif not selected_source_of_funds:
            if is_ajax:
                return JsonResponse({'success': False, 'errors': {'source_of_funds': ['Source of Funds is required.']}}, status=400)
            from django.contrib import messages
            messages.error(request, 'Source of Funds is required.')
            return redirect('project_list')

        # Allow "Add new project type" (ProjectType FK requires an ID)
        new_project_type = (post_data.get('project_type_new') or '').strip()
        selected_project_type = (post_data.get('project_type') or '').strip()
        # Treat __add_new__ or empty as "use project_type_new" when user typed a new type
        if new_project_type and (not selected_project_type or selected_project_type == '__add_new__'):
            pt = _get_or_create_project_type_by_name(new_project_type)
            if pt:
                post_data['project_type'] = str(pt.id)
        
        # Check if this is an edit operation
        project_id = request.POST.get('project_id')
        is_edit = bool(project_id)
        
        if is_edit:
            # Get existing project for editing
            try:
                existing_project = Project.objects.get(pk=project_id)
                form = ProjectForm(post_data, request.FILES, instance=existing_project)
            except Project.DoesNotExist:
                from django.http import JsonResponse
                if is_ajax:
                    return JsonResponse({'success': False, 'error': 'Project not found'}, status=404)
                from django.contrib import messages
                messages.error(request, 'Project not found.')
                return redirect('project_list')
        else:
            form = ProjectForm(post_data, request.FILES)
        
        if form.is_valid():
            project = form.save(commit=False)
            if not is_edit:
                project.created_by = request.user
            
            # Phase 4: Auto-detect zone before saving
            # Try to detect zone, but don't block project creation if it fails
            zone_type, confidence = None, 0
            try:
                # Try zone detection if we have barangay, name, or description
                if project.barangay or project.name or project.description:
                    zone_type, confidence = project.detect_and_set_zone(save=False)
                    if zone_type:
                        # Set the zone_type on the project object
                        project.zone_type = zone_type
                        # Auto-validate if confidence is high
                        if confidence >= 70:
                            project.zone_validated = True
            except Exception as e:
                # Fail gracefully - don't block project creation if zone detection fails
                import logging
                logger = logging.getLogger(__name__)
                logger.warning(f"Zone detection failed for project: {str(e)}")
                zone_type, confidence = None, 0
            
            # Log image upload info for debugging (use print() so it shows in runtime logs)
            if 'image' in request.FILES:
                print(f"📤 Uploading image: {request.FILES['image'].name}, size: {request.FILES['image'].size} bytes")
                print(f"   Content type: {request.FILES['image'].content_type}")
            
            try:
                # Test storage connection before saving
                from django.core.files.storage import default_storage
                from django.conf import settings
                
                print(f"Storage backend: {default_storage.__class__.__name__}")
                if hasattr(settings, 'DEFAULT_FILE_STORAGE'):
                    print(f"DEFAULT_FILE_STORAGE: {settings.DEFAULT_FILE_STORAGE}")
                
                # Save the project (this should trigger file upload to Spaces)
                print("Saving project...")
                project.save()
                print("Project saved, saving many-to-many relationships...")
                form.save_m2m()  # Save assigned engineers
                print("✅ Project and relationships saved successfully!")
                
                # Log successful save and image URL
                if project.image:
                    print(f"   Image field value: {project.image}")
                    print(f"   Image name: {project.image.name}")
                    try:
                        image_url = project.image.url
                        print(f"   Image URL: {image_url}")
                        
                        # Try to verify the file exists in storage
                        print(f"   Checking if file exists in storage: {project.image.name}")
                        if default_storage.exists(project.image.name):
                            print(f"   ✅ File exists in storage: {project.image.name}")
                            
                            # Try to get file size
                            try:
                                file_size = default_storage.size(project.image.name)
                                print(f"   File size: {file_size} bytes")
                            except Exception as size_error:
                                print(f"   ⚠️  Could not get file size: {str(size_error)}")
                        else:
                            print(f"   ❌ File does NOT exist in storage: {project.image.name}")
                            print(f"   This means the upload to Spaces failed!")
                            
                            # Try to check storage connection
                            try:
                                print("   Testing storage connection...")
                                # Try to list files (this will test the connection)
                                print(f"   Storage class: {default_storage.__class__}")
                            except Exception as conn_error:
                                print(f"   ❌ Storage connection error: {str(conn_error)}")
                                import traceback
                                print(f"   Traceback: {traceback.format_exc()}")
                    except Exception as url_error:
                        print(f"   ❌ Error getting image URL: {str(url_error)}")
                        import traceback
                        print(f"   Traceback: {traceback.format_exc()}")
                else:
                    print("⚠️  No image field set after save")
            except Exception as e:
                print(f"❌ Error saving project: {str(e)}")
                import traceback
                print(f"Full traceback: {traceback.format_exc()}")
                
                # Return JSON error for AJAX requests
                if is_ajax:
                    from django.http import JsonResponse
                    return JsonResponse({'success': False, 'error': f'Error saving project: {str(e)}'}, status=400)
                raise
            
            # Log zone detection result (optional, for debugging)
            if zone_type:
                import logging
                logger = logging.getLogger(__name__)
                logger.info(f"Zone detected for project '{project.name}': {zone_type} (confidence: {confidence}%)")
            
            # Return JSON response for AJAX requests
            if is_ajax:
                from django.http import JsonResponse
                message = 'Project updated successfully!' if is_edit else 'Project created successfully!'
                return JsonResponse({'success': True, 'message': message, 'project_id': project.id})
            
            # Success message including assigned engineers so user can verify
            from django.contrib import messages
            assigned = list(project.assigned_engineers.values_list('username', flat=True))
            assigned_str = ', '.join(assigned) if assigned else 'None'
            if is_edit:
                messages.success(request, f'Project updated successfully! Assigned engineers: {assigned_str}.')
            else:
                messages.success(request, f'Project created successfully! Assigned engineers: {assigned_str}.')
            return redirect('project_list')
        else:
            # Return JSON errors for AJAX requests
            if is_ajax:
                from django.http import JsonResponse
                errors = {}
                for field, error_list in form.errors.items():
                    errors[field] = error_list
                return JsonResponse({'success': False, 'errors': errors, 'error': 'Please correct the errors below.'}, status=400)
            
            # Re-render with errors for non-AJAX requests
            if is_head_engineer(request.user) or is_finance_manager(request.user):
                projects = Project.objects.all().select_related('project_type').order_by('-created_at')
            elif is_project_engineer(request.user):
                projects = Project.objects.filter(assigned_engineers=request.user).select_related('project_type').order_by('-created_at')
            else:
                projects = Project.objects.none()
            paginator = Paginator(projects, 10)
            page_number = request.GET.get('page')
            page_obj = paginator.get_page(page_number)
            return render(request, 'monitoring/project_list.html', {'page_obj': page_obj, 'form': form})
    # GET logic
    if is_head_engineer(request.user) or is_finance_manager(request.user):
        projects = Project.objects.all().select_related('project_type')
    elif is_project_engineer(request.user):
        projects = Project.objects.filter(assigned_engineers=request.user).select_related('project_type')
    else:
        projects = Project.objects.none()

    # Ensure SourceOfFunds master list is seeded from existing projects (one-time)
    if not SourceOfFunds.objects.exists():
        existing_sources = (
            Project.objects.exclude(source_of_funds__isnull=True)
            .exclude(source_of_funds='')
            .values_list('source_of_funds', flat=True)
            .distinct()
        )
        for name in existing_sources:
            cleaned = (name or '').strip()
            if cleaned and not SourceOfFunds.objects.filter(name__iexact=cleaned).exists():
                SourceOfFunds.objects.create(name=cleaned)
    source_of_funds_options = SourceOfFunds.objects.filter(is_active=True).order_by('name')
    project_type_options = ProjectType.objects.all().order_by('name')
    
    # Apply filters
    from django.db.models import Q
    from decimal import Decimal, InvalidOperation
    barangay_filter = request.GET.get('barangay')
    duration_filter = request.GET.get('duration')
    status_filter = request.GET.get('status')
    source_of_funds_filter = (request.GET.get('source_of_funds') or '').strip()
    project_type_filter = (request.GET.get('project_type') or '').strip()
    cost_min_raw = (request.GET.get('cost_min') or '').strip()
    cost_max_raw = (request.GET.get('cost_max') or '').strip()
    search_query = request.GET.get('search', '').strip()
    
    # Barangay filter
    if barangay_filter:
        projects = projects.filter(barangay=barangay_filter)
    
    # Status filter
    if status_filter:
        # Handle "in_progress" to match both "in_progress" and "ongoing" statuses
        if status_filter == 'in_progress':
            projects = projects.filter(Q(status='in_progress') | Q(status='ongoing'))
        elif status_filter == 'delayed':
            # For delayed, we need to filter projects that are either:
            # 1. Already marked as delayed in database, OR
            # 2. Should be delayed (in_progress/ongoing + end_date passed + progress < 99%)
            # We'll filter these in Python after getting progress data
            delayed_marked = projects.filter(status='delayed')
            potentially_delayed = projects.filter(Q(status='in_progress') | Q(status='ongoing'))
            projects = delayed_marked | potentially_delayed
        else:
            projects = projects.filter(status=status_filter)

    # Source of Funds filter (case-insensitive exact match)
    if source_of_funds_filter:
        projects = projects.filter(source_of_funds__iexact=source_of_funds_filter)

    # Project type filter
    if project_type_filter:
        try:
            projects = projects.filter(project_type_id=int(project_type_filter))
        except (TypeError, ValueError):
            pass

    # Project cost min/max filters
    if cost_min_raw:
        try:
            projects = projects.filter(project_cost__gte=Decimal(cost_min_raw))
        except (InvalidOperation, ValueError):
            pass
    if cost_max_raw:
        try:
            projects = projects.filter(project_cost__lte=Decimal(cost_max_raw))
        except (InvalidOperation, ValueError):
            pass
    
    # Search filter (search in name, PRN, description)
    if search_query:
        projects = projects.filter(
            Q(name__icontains=search_query) |
            Q(prn__icontains=search_query) |
            Q(description__icontains=search_query)
        )
    
    # Duration filter (based on project duration calculated from start_date and end_date)
    if duration_filter:
        # Filter projects that have both start and end dates, then filter by duration in Python
        # This is more reliable than raw SQL and works across all databases
        projects_with_dates = projects.filter(
            start_date__isnull=False,
            end_date__isnull=False
        )
        
        # Get project IDs that match the duration criteria
        matching_project_ids = []
        for project in projects_with_dates.only('id', 'start_date', 'end_date'):
            if project.start_date and project.end_date:
                duration_days = (project.end_date - project.start_date).days
                
                if duration_filter == 'lt6' and duration_days < 180:
                    matching_project_ids.append(project.id)
                elif duration_filter == '6to12' and 180 <= duration_days <= 365:
                    matching_project_ids.append(project.id)
                elif duration_filter == 'gt12' and duration_days > 365:
                    matching_project_ids.append(project.id)
        
        # Filter the queryset to only include matching projects
        if matching_project_ids:
            projects = projects.filter(id__in=matching_project_ids)
        else:
            # No projects match the duration criteria
            projects = projects.none()
    
    # Order by created_at descending
    projects = projects.order_by('-created_at')
    
    # Get latest progress for all projects to calculate delayed status
    from django.utils import timezone
    from django.db.models import Max
    from projeng.models import ProjectProgress
    project_ids = [p.id for p in projects]
    latest_progress = {}
    if project_ids:
        latest_progress_qs = ProjectProgress.objects.filter(
            project_id__in=project_ids
        ).values('project_id').annotate(
            latest_date=Max('date'),
            latest_created=Max('created_at')
        )
        for item in latest_progress_qs:
            latest = ProjectProgress.objects.filter(
                project_id=item['project_id'],
                date=item['latest_date'],
                created_at=item['latest_created']
            ).order_by('-created_at').first()
            if latest and latest.percentage_complete is not None:
                latest_progress[item['project_id']] = int(latest.percentage_complete)
    
    today = timezone.now().date()
    
    # If filtering by delayed, filter the queryset to only include delayed projects
    if status_filter == 'delayed':
        delayed_project_ids = []
        for p in projects:
            progress = latest_progress.get(p.id, 0)
            stored_status = p.status or ''
            # Check if project should be delayed
            if stored_status == 'delayed':
                delayed_project_ids.append(p.id)
            elif progress < 99 and p.end_date and p.end_date < today and stored_status in ['in_progress', 'ongoing']:
                delayed_project_ids.append(p.id)
        projects = projects.filter(id__in=delayed_project_ids)
    
    paginator = Paginator(projects, 15)  # Show 15 projects per page
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    form = ProjectForm()
    # Build projects_data for modal and JS (use filtered projects)
    projects_data = []
    import logging
    logger = logging.getLogger(__name__)

    # Pre-calculate spent/remaining for current page projects
    from django.db.models import Sum
    from decimal import Decimal
    page_project_ids = [p.id for p in page_obj.object_list]
    spent_by_project_id = {}
    if page_project_ids:
        for row in (
            ProjectCost.objects.filter(project_id__in=page_project_ids)
            .values('project_id')
            .annotate(total_spent=Sum('amount'))
        ):
            spent_by_project_id[row['project_id']] = row['total_spent'] or Decimal('0.00')
    
    for p in page_obj.object_list:
        try:
            # Get progress
            progress = latest_progress.get(p.id, 0)
            if progress == 0:
                progress = getattr(p, 'progress', 0) or 0
            
            # Calculate actual status dynamically (including delayed)
            stored_status = p.status or ''
            calculated_status = stored_status
            
            # Priority: completed > delayed > in_progress > planned
            if progress >= 99:
                calculated_status = 'completed'
            elif stored_status == 'delayed':
                calculated_status = 'delayed'
            elif progress < 99 and p.end_date and p.end_date < today and stored_status in ['in_progress', 'ongoing']:
                # Project is delayed if: end_date passed, progress < 99%, and status is in_progress/ongoing
                calculated_status = 'delayed'
            elif stored_status in ['in_progress', 'ongoing']:
                calculated_status = 'in_progress'
            elif stored_status in ['planned', 'pending']:
                calculated_status = 'planned'
            
            # Safely get image URL
            image_url = ''
            if p.image:
                try:
                    image_url = p.image.url
                except Exception as img_error:
                    logger.warning(f"Error getting image URL for project {p.id}: {str(img_error)}")
                    image_url = ''
            
            # Safely get assigned engineers
            assigned_engineers = []
            try:
                if hasattr(p, 'assigned_engineers'):
                    assigned_engineers = [str(e) for e in p.assigned_engineers.all()]
            except Exception as eng_error:
                logger.warning(f"Error getting assigned engineers for project {p.id}: {str(eng_error)}")
                assigned_engineers = []
            
            # Auto-detect zone type if not assigned
            zone_type = p.zone_type or ''
            if not zone_type and (p.barangay or p.name or p.description):
                try:
                    detected_zone, confidence = p.detect_and_set_zone(save=True)
                    if detected_zone:
                        zone_type = detected_zone
                        logger.info(f"Auto-detected zone for project {p.id}: {zone_type} (confidence: {confidence}%)")
                except Exception as zone_error:
                    logger.warning(f"Zone detection failed for project {p.id}: {str(zone_error)}")
                    # Continue without zone type if detection fails
            
            budget_val = p.project_cost if p.project_cost is not None else None
            spent_val = spent_by_project_id.get(p.id, Decimal('0.00'))
            remaining_val = (budget_val - spent_val) if budget_val is not None else None

            projects_data.append({
                'id': p.id,
                'name': p.name or '',
                'prn': p.prn or '',
                'description': p.description or '',
                'barangay': p.barangay or '',
                'latitude': float(p.latitude) if p.latitude else '',
                'longitude': float(p.longitude) if p.longitude else '',
                'project_cost': f"{budget_val:.2f}" if budget_val is not None else '',
                'total_spent': f"{spent_val:.2f}",
                'remaining_funds': f"{remaining_val:.2f}" if remaining_val is not None else '',
                'source_of_funds': p.source_of_funds or '',
                'start_date': str(p.start_date) if p.start_date else '',
                'end_date': str(p.end_date) if p.end_date else '',
                'status': calculated_status,  # Use calculated_status instead of stored status
                'image': image_url,
                'progress': progress,
                'assigned_engineers': assigned_engineers,
                'zone_type': zone_type,
            })
        except Exception as e:
            logger.error(f"Error building project data for project {p.id}: {str(e)}", exc_info=True)
            # Continue with other projects even if one fails
            continue
    # Ensure projects_data is a list (not a string or other type)
    if not isinstance(projects_data, list):
        logger.warning(f"projects_data is not a list, converting. Type: {type(projects_data)}")
        projects_data = list(projects_data) if projects_data else []
    
    # Create a dictionary mapping project IDs to calculated statuses for template use
    calculated_statuses = {}
    for p_data in projects_data:
        calculated_statuses[p_data['id']] = p_data['status']
    
    # Update project objects in page_obj with calculated status for template display
    for project in page_obj.object_list:
        if project.id in calculated_statuses:
            # Temporarily set the status attribute for template display
            project.calculated_status = calculated_statuses[project.id]
        else:
            project.calculated_status = project.status or ''
    
    # Log how many projects we're sending to the template
    logger.info(f"Sending {len(projects_data)} projects to template")
    
    return render(request, 'monitoring/project_list.html', {
        'page_obj': page_obj,
        'form': form,
        'projects_data': projects_data,
        'status_filter': status_filter,
        'barangay_filter': barangay_filter,
        'duration_filter': duration_filter,
        'source_of_funds_filter': source_of_funds_filter,
        'source_of_funds_options': source_of_funds_options,
        'project_type_options': project_type_options,
        'project_type_filter': project_type_filter,
        'cost_min': cost_min_raw,
        'cost_max': cost_max_raw,
    })


@login_required
def create_project_type_api(request):
    """Create a new ProjectType (or return existing) via AJAX."""
    if request.method != 'POST':
        return JsonResponse({"success": False, "error": "Method not allowed"}, status=405)
    if not (is_head_engineer(request.user) or is_finance_manager(request.user) or request.user.is_superuser):
        return JsonResponse({"success": False, "error": "Not allowed"}, status=403)

    try:
        payload = json.loads((request.body or b"{}").decode("utf-8"))
    except Exception:
        return JsonResponse({"success": False, "error": "Invalid JSON body"}, status=400)

    name = (payload.get("name") or "").strip()
    if not name:
        return JsonResponse({"success": False, "error": "Project type name is required"}, status=400)

    pt = _get_or_create_project_type_by_name(name)
    if not pt:
        return JsonResponse({"success": False, "error": "Could not create project type"}, status=400)

    return JsonResponse({"success": True, "id": pt.id, "name": pt.name, "code": pt.code})


@login_required
def generate_prn_api(request):
    """
    Generate a unique PRN for new projects.
    Used by the Add Project modal so PRN is visible but not editable.
    """
    if request.method != 'GET':
        return JsonResponse({"success": False, "error": "Method not allowed"}, status=405)

    if not (is_head_engineer(request.user) or is_finance_manager(request.user) or request.user.is_superuser):
        return JsonResponse({"success": False, "error": "Not allowed"}, status=403)

    from django.utils import timezone
    import secrets
    from projeng.models import Project

    date_part = timezone.now().strftime("%y%m%d")
    for _ in range(50):
        rand_part = secrets.token_hex(3).upper()
        candidate = f"PRN-{date_part}-{rand_part}"
        if not Project.objects.filter(prn=candidate).exists():
            return JsonResponse({"success": True, "prn": candidate})

    return JsonResponse({"success": False, "error": "Could not generate PRN"}, status=500)

@login_required
@prevent_project_engineer_access
def map_view(request):
    try:
        from projeng.models import SourceOfFunds, ProjectType
        if is_head_engineer(request.user) or is_finance_manager(request.user):
            projects = Project.objects.all()
        elif is_project_engineer(request.user):
            projects = Project.objects.filter(assigned_engineers=request.user)
        else:
            projects = Project.objects.none()
        # Only include projects with coordinates
        projects_with_coords = projects.filter(latitude__isnull=False, longitude__isnull=False).prefetch_related('assigned_engineers')

        # Seed SourceOfFunds master list if empty (so dropdown isn't blank)
        if not SourceOfFunds.objects.exists():
            existing_sources = (
                Project.objects.exclude(source_of_funds__isnull=True)
                .exclude(source_of_funds='')
                .values_list('source_of_funds', flat=True)
                .distinct()
            )
            for name in existing_sources:
                cleaned = (name or '').strip()
                if cleaned and not SourceOfFunds.objects.filter(name__iexact=cleaned).exists():
                    SourceOfFunds.objects.create(name=cleaned)
        
        # Get latest progress for all projects
        from django.db.models import Max
        from django.utils import timezone
        project_ids = [p.id for p in projects_with_coords]
        latest_progress = {}
        if project_ids:
            from projeng.models import ProjectProgress
            latest_progress_qs = ProjectProgress.objects.filter(
                project_id__in=project_ids
            ).values('project_id').annotate(
                latest_date=Max('date'),
                latest_created=Max('created_at')
            )
            for item in latest_progress_qs:
                latest = ProjectProgress.objects.filter(
                    project_id=item['project_id'],
                    date=item['latest_date'],
                    created_at=item['latest_created']
                ).order_by('-created_at').first()
                if latest and latest.percentage_complete is not None:
                    latest_progress[item['project_id']] = int(latest.percentage_complete)
        
        today = timezone.now().date()
        projects_data = []
        # Pre-calculate spent/remaining for projects in map
        from django.db.models import Sum
        from decimal import Decimal
        map_project_ids = list(projects_with_coords.values_list('id', flat=True))
        spent_by_project_id = {}
        if map_project_ids:
            for row in (
                ProjectCost.objects.filter(project_id__in=map_project_ids)
                .values('project_id')
                .annotate(total_spent=Sum('amount'))
            ):
                spent_by_project_id[row['project_id']] = row['total_spent'] or Decimal('0.00')

        for p in projects_with_coords:
            try:
                # Get assigned engineers as a list of usernames
                assigned_engineers = [eng.username for eng in p.assigned_engineers.all()]
                # Get progress - prefer latest progress update, fallback to project.progress field
                progress_value = latest_progress.get(p.id)
                if progress_value is None:
                    progress_value = getattr(p, 'progress', 0)
                    if progress_value is None:
                        progress_value = 0
                
                # Calculate actual status dynamically (including delayed)
                calculated_status = p.status or ''
                if progress_value >= 99:
                    calculated_status = 'completed'
                elif progress_value < 99 and p.end_date and p.end_date < today and p.status in ['in_progress', 'ongoing']:
                    # Project is delayed if: end_date passed, progress < 99%, and status is in_progress/ongoing
                    calculated_status = 'delayed'
                elif p.status in ['in_progress', 'ongoing']:
                    calculated_status = 'in_progress'
                elif p.status in ['planned', 'pending']:
                    calculated_status = 'planned'
                
                # Safely get image URL
                image_url = ""
                try:
                    if p.image:
                        image_url = p.image.url
                except (ValueError, AttributeError):
                    image_url = ""
                
                # Auto-detect zone type if not assigned
                zone_type = p.zone_type or ''
                if not zone_type and (p.barangay or p.name or p.description):
                    try:
                        detected_zone, confidence = p.detect_and_set_zone(save=True)
                        if detected_zone:
                            zone_type = detected_zone
                            import logging
                            logger = logging.getLogger(__name__)
                            logger.info(f"Auto-detected zone for project {p.id} in map_view: {zone_type} (confidence: {confidence}%)")
                    except Exception as zone_error:
                        import logging
                        logger = logging.getLogger(__name__)
                        logger.warning(f"Zone detection failed for project {p.id} in map_view: {str(zone_error)}")
                        # Continue without zone type if detection fails
                
                budget_val = p.project_cost if p.project_cost is not None else None
                spent_val = spent_by_project_id.get(p.id, Decimal('0.00'))
                remaining_val = (budget_val - spent_val) if budget_val is not None else None

                projects_data.append({
                    'id': p.id,
                    'name': p.name or '',
                    'latitude': float(p.latitude) if p.latitude else 0.0,
                    'longitude': float(p.longitude) if p.longitude else 0.0,
                    'barangay': p.barangay or '',
                    'status': calculated_status,  # Use calculated_status instead of p.status
                    'description': p.description or '',
                    'project_cost': f"{budget_val:.2f}" if budget_val is not None else "",
                    'total_spent': f"{spent_val:.2f}",
                    'remaining_funds': f"{remaining_val:.2f}" if remaining_val is not None else "",
                    'source_of_funds': p.source_of_funds or '',
                    'prn': p.prn or '',
                    'start_date': str(p.start_date) if p.start_date else "",
                    'end_date': str(p.end_date) if p.end_date else "",
                    'image': image_url,
                    'progress': progress_value,
                    'assigned_engineers': assigned_engineers,
                    'zone_type': zone_type,
                })
            except Exception as e:
                import logging
                logger = logging.getLogger(__name__)
                logger.error(f'Error processing project {p.id} in map_view: {str(e)}', exc_info=True)
                continue
        
        # Pass user context for template permission checks
        # is_head_engineer is already imported at the top of the file
        context = {
            'projects_data': projects_data,
            'is_head_engineer': is_head_engineer(request.user),
            'source_of_funds_options': SourceOfFunds.objects.filter(is_active=True).order_by('name'),
            'project_type_options': ProjectType.objects.all().order_by('name'),
        }
        return render(request, 'monitoring/map.html', context)
    except Exception as e:
        import logging
        logger = logging.getLogger(__name__)
        logger.error(f'Error in map_view: {str(e)}', exc_info=True)
        from django.http import HttpResponseServerError
        return HttpResponseServerError('An error occurred while loading the map view.')

@login_required
@csrf_exempt
def overall_project_metrics_api(request):
    """API endpoint to fetch overall project metrics - accessible to all authenticated users"""
    try:
        from projeng.models import Project, ProjectProgress
        from django.db.models import Max
        from django.utils import timezone
        import logging
        
        logger = logging.getLogger(__name__)
        
        # Role-based queryset - get ALL projects (not just those with coordinates)
        if is_head_engineer(request.user) or is_finance_manager(request.user):
            projects_qs = Project.objects.all()
            logger.info(f'User {request.user.username} is head engineer or finance manager, fetching all projects')
        elif is_project_engineer(request.user):
            projects_qs = Project.objects.filter(assigned_engineers=request.user)
            logger.info(f'User {request.user.username} is project engineer, fetching assigned projects')
        else:
            projects_qs = Project.objects.none()
            logger.warning(f'User {request.user.username} has no recognized role, returning empty queryset')
        
        # Convert queryset to list to ensure consistent evaluation
        projects_list = list(projects_qs)
        project_count = len(projects_list)
        logger.info(f'Found {project_count} projects for user {request.user.username}')
        
        # Get latest progress for all projects
        project_ids = [p.id for p in projects_list]
        latest_progress = {}
        if project_ids:
            latest_progress_qs = ProjectProgress.objects.filter(
                project_id__in=project_ids
            ).values('project_id').annotate(
                latest_date=Max('date'),
                latest_created=Max('created_at')
            )
            for item in latest_progress_qs:
                latest = ProjectProgress.objects.filter(
                    project_id=item['project_id'],
                    date=item['latest_date'],
                    created_at=item['latest_created']
                ).order_by('-created_at').first()
                if latest and latest.percentage_complete is not None:
                    latest_progress[item['project_id']] = int(latest.percentage_complete)
        
        # Calculate status counts dynamically
        today = timezone.now().date()
        total_projects = 0
        completed_count = 0
        in_progress_count = 0
        planned_count = 0
        delayed_count = 0
        
        for p in projects_list:
            total_projects += 1
            progress = latest_progress.get(p.id, 0)
            stored_status = (p.status or '').lower().strip()
            
            # Calculate actual status dynamically
            # Priority: completed > delayed > in_progress > planned
            # Check for completed: either progress >= 99 OR stored status is 'completed'
            if progress >= 99 or stored_status == 'completed':
                completed_count += 1
                logger.debug(f'Project {p.id} ({p.name}): COMPLETED - progress={progress}%, status="{p.status}"')
            elif stored_status == 'delayed':
                delayed_count += 1
                logger.debug(f'Project {p.id} ({p.name}): DELAYED - progress={progress}%, status="{p.status}"')
            elif progress < 99 and p.end_date and p.end_date < today and stored_status in ['in_progress', 'ongoing']:
                # Project is delayed if: end_date passed, progress < 99%, and status is in_progress/ongoing
                delayed_count += 1
                logger.debug(f'Project {p.id} ({p.name}): DELAYED (overdue) - progress={progress}%, status="{p.status}", end_date={p.end_date}')
            elif stored_status in ['in_progress', 'ongoing']:
                in_progress_count += 1
                logger.debug(f'Project {p.id} ({p.name}): IN PROGRESS - progress={progress}%, status="{p.status}"')
            elif stored_status in ['planned', 'pending']:
                planned_count += 1
                logger.debug(f'Project {p.id} ({p.name}): PLANNED - progress={progress}%, status="{p.status}"')
            else:
                # If status doesn't match any category, default to planned or log it
                logger.warning(f'Project {p.id} ({p.name}) has unrecognized status: "{p.status}", defaulting to planned. Progress: {progress}%')
                planned_count += 1
        
        metrics = {
            'total_projects': total_projects,
            'completed': completed_count,
            'in_progress': in_progress_count,
            'planned': planned_count,
            'delayed': delayed_count
        }
        
        logger.info(f'Metrics calculated for user {request.user.username}: {metrics}')
        
        return JsonResponse({
            'success': True,
            'metrics': metrics
        })
    except Exception as e:
        import logging
        import traceback
        logger = logging.getLogger(__name__)
        error_trace = traceback.format_exc()
        logger.error(f'Error in overall_project_metrics_api: {str(e)}\n{error_trace}')
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=500)

@login_required
@prevent_project_engineer_access
def reports(request):
    # This view is only accessible to Head Engineers and Finance Managers
    # Project Engineers are redirected by the decorator
    from projeng.models import ProjectCost, SourceOfFunds
    from django.db.models import Q
    
    if is_head_engineer(request.user) or is_finance_manager(request.user):
        projects = Project.objects.all()
    else:
        projects = Project.objects.none()
    
    # Apply filters from request parameters
    barangay_filter = request.GET.get('barangay', '').strip()
    status_filter = request.GET.get('status', '').strip()
    start_date_filter = request.GET.get('start_date', '').strip()
    end_date_filter = request.GET.get('end_date', '').strip()
    source_of_funds_filter = request.GET.get('source_of_funds', '').strip()
    cost_min_filter = request.GET.get('cost_min', '').strip()
    cost_max_filter = request.GET.get('cost_max', '').strip()
    
    # Filter by barangay
    if barangay_filter:
        projects = projects.filter(barangay=barangay_filter)
    
    # Filter by status
    if status_filter:
        # Handle "in_progress" to match both "in_progress" and "ongoing" statuses
        if status_filter == 'in_progress':
            projects = projects.filter(Q(status='in_progress') | Q(status='ongoing'))
        else:
            projects = projects.filter(status=status_filter)
    
    # Filter by start date
    if start_date_filter:
        try:
            from datetime import datetime
            start_date = datetime.strptime(start_date_filter, '%Y-%m-%d').date()
            projects = projects.filter(start_date__gte=start_date)
        except ValueError:
            pass
    
    # Filter by end date
    if end_date_filter:
        try:
            from datetime import datetime
            end_date = datetime.strptime(end_date_filter, '%Y-%m-%d').date()
            projects = projects.filter(end_date__lte=end_date)
        except ValueError:
            pass
    
    # Filter by source of funds
    if source_of_funds_filter:
        projects = projects.filter(source_of_funds__iexact=source_of_funds_filter)
    
    # Filter by project cost range
    if cost_min_filter:
        try:
            cost_min = float(cost_min_filter)
            projects = projects.filter(project_cost__gte=cost_min)
        except (ValueError, TypeError):
            pass
    if cost_max_filter:
        try:
            cost_max = float(cost_max_filter)
            projects = projects.filter(project_cost__lte=cost_max)
        except (ValueError, TypeError):
            pass
    
    # Calculate budget metrics
    total_budget = 0
    total_spent = 0
    projects_with_budget = 0
    
    for p in projects:
        if p.project_cost:
            try:
                total_budget += float(p.project_cost)
                projects_with_budget += 1
            except (ValueError, TypeError):
                pass
    
    # Calculate total spent from ProjectCost
    project_ids = [p.id for p in projects]
    if project_ids:
        costs = ProjectCost.objects.filter(project_id__in=project_ids)
        for cost in costs:
            try:
                if cost.amount:
                    total_spent += float(cost.amount)
            except (ValueError, TypeError):
                pass
    
    remaining_budget = total_budget - total_spent
    budget_utilization = (total_spent / total_budget * 100) if total_budget > 0 else 0
    avg_project_cost = (total_budget / len(projects)) if len(projects) > 0 else 0

    # Get latest progress per project
    from projeng.models import ProjectProgress
    from django.db.models import Max
    latest_progress_map = {}
    if project_ids:
        progress_qs = ProjectProgress.objects.filter(project_id__in=project_ids).values('project_id').annotate(
            latest_date=Max('date'), latest_created=Max('created_at')
        )
        for item in progress_qs:
            latest = ProjectProgress.objects.filter(
                project_id=item['project_id'], date=item['latest_date'], created_at=item['latest_created']
            ).order_by('-id').first()
            if latest and latest.percentage_complete is not None:
                latest_progress_map[item['project_id']] = float(latest.percentage_complete)

    # Prepare data for JS and summary cards
    projects_list = []
    status_map = defaultdict(int)
    barangay_map = defaultdict(int)
    for p in projects:
        progress_val = latest_progress_map.get(p.id)
        if progress_val is None:
            progress_val = getattr(p, 'progress', 0) or 0
        projects_list.append({
            'id': p.id,
            'name': p.name,
            'prn': p.prn,
            'description': p.description,
            'barangay': p.barangay,
            'location': p.barangay,  # For compatibility
            'project_cost': str(p.project_cost) if p.project_cost is not None else '',
            'source_of_funds': p.source_of_funds,
            'start_date': str(p.start_date) if p.start_date else '',
            'end_date': str(p.end_date) if p.end_date else '',
            'status': p.status,
            'progress': progress_val,
        })
        # For charts
        status = p.status
        if status in ['in_progress', 'ongoing']:
            status_map['Ongoing'] += 1
        elif status in ['planned', 'pending']:
            status_map['Planned'] += 1
        elif status == 'completed':
            status_map['Completed'] += 1
        elif status == 'delayed':
            status_map['Delayed'] += 1
        else:
            status_map[status.title()] += 1
        if p.barangay:
            barangay_map[p.barangay] += 1
    status_labels = list(status_map.keys())
    status_counts = list(status_map.values())
    barangay_labels = list(barangay_map.keys())
    barangay_counts = list(barangay_map.values())
    context = {
        'status_labels': json.dumps(status_labels),
        'status_counts': json.dumps(status_counts),
        'barangay_labels': json.dumps(barangay_labels),
        'barangay_counts': json.dumps(barangay_counts),
        'projects_json': projects_list,  # Pass list directly, let json_script handle encoding
        'projects_count': len(projects_list),
        'total_budget': total_budget,
        'total_spent': total_spent,
        'remaining_budget': remaining_budget,
        'budget_utilization': budget_utilization,
        'avg_project_cost': avg_project_cost,
        'selected_barangay': barangay_filter,
        'selected_status': status_filter,
        'selected_start_date': start_date_filter,
        'selected_end_date': end_date_filter,
        'source_of_funds_options': SourceOfFunds.objects.filter(is_active=True).order_by('name'),
        'selected_source_of_funds': source_of_funds_filter,
        'selected_cost_min': cost_min_filter,
        'selected_cost_max': cost_max_filter,
    }
    return render(request, 'monitoring/reports.html', context)

@login_required
@prevent_project_engineer_access
def budget_reports(request):
    from projeng.models import Project, ProjectCost
    from collections import defaultdict
    from django.db.models import Q
    from django.core.paginator import Paginator
    
    # Role-based queryset
    if is_head_engineer(request.user) or is_finance_manager(request.user):
        projects = Project.objects.all()
    elif is_project_engineer(request.user):
        projects = Project.objects.filter(assigned_engineers=request.user)
    else:
        projects = Project.objects.none()
    
    # Apply filters
    selected_barangay = request.GET.get('barangay', '').strip()
    selected_status = request.GET.get('status', '').strip()
    selected_start_date = request.GET.get('start_date', '').strip()
    selected_end_date = request.GET.get('end_date', '').strip()
    selected_cost_min = request.GET.get('cost_min', '').strip()
    selected_cost_max = request.GET.get('cost_max', '').strip()
    selected_budget_status = request.GET.get('budget_status', '').strip()
    selected_chart_period = request.GET.get('chart_period', 'month').strip().lower()
    if selected_chart_period not in ('week', 'month', 'year'):
        selected_chart_period = 'month'
    
    if selected_barangay:
        projects = projects.filter(barangay=selected_barangay)
    
    if selected_status:
        if selected_status == 'in_progress':
            projects = projects.filter(Q(status='in_progress') | Q(status='ongoing'))
        else:
            projects = projects.filter(status=selected_status)
    
    if selected_start_date:
        projects = projects.filter(start_date__gte=selected_start_date)
    
    if selected_end_date:
        projects = projects.filter(end_date__lte=selected_end_date)

    # Cost range filters (by project budget)
    if selected_cost_min:
        try:
            cost_min_val = float(selected_cost_min)
            projects = projects.filter(project_cost__gte=cost_min_val)
        except (ValueError, TypeError):
            pass
    if selected_cost_max:
        try:
            cost_max_val = float(selected_cost_max)
            projects = projects.filter(project_cost__lte=cost_max_val)
        except (ValueError, TypeError):
            pass
    
    # Get all projects for summary calculations (before pagination)
    all_projects_for_summary = projects
    
    # Chart period filter: projects overlapping current week/month/year
    from datetime import date, timedelta
    today = timezone.now().date() if timezone.is_naive(timezone.now()) else timezone.now().date()
    if selected_chart_period == 'week':
        period_start = today - timedelta(days=today.weekday())
        period_end = period_start + timedelta(days=6)
    elif selected_chart_period == 'year':
        period_start = date(today.year, 1, 1)
        period_end = date(today.year, 12, 31)
    else:
        period_start = date(today.year, today.month, 1)
        if today.month == 12:
            period_end = date(today.year, 12, 31)
        else:
            period_end = date(today.year, today.month + 1, 1) - timedelta(days=1)
    
    def project_in_chart_period(p):
        if not getattr(p, 'start_date', None) and not getattr(p, 'end_date', None):
            return True
        start = getattr(p, 'start_date', None)
        end = getattr(p, 'end_date', None)
        if start and end:
            return start <= period_end and end >= period_start
        if start:
            return start <= period_end
        if end:
            return end >= period_start
        return True
    
    projects_for_chart = [p for p in all_projects_for_summary if project_in_chart_period(p)]
    
    project_data = []
    project_names = []
    utilizations = []
    over_count = 0
    under_count = 0
    within_count = 0
    total_budget_sum = 0
    total_spent_sum = 0
    at_risk_count = 0
    
    # Calculate summary from all filtered projects
    for p in all_projects_for_summary:
        # Calculate spent and utilization
        costs = ProjectCost.objects.filter(project=p)
        spent = sum([float(c.amount) for c in costs]) if costs else 0
        budget = float(p.project_cost) if p.project_cost else 0
        remaining = budget - spent
        utilization = (spent / budget * 100) if budget > 0 else 0
        
        # Determine budget status using same logic as finance_cost_management
        if budget > 0:
            if spent > budget:
                over_under = 'Over'
            elif utilization >= 80:
                over_under = 'Within'
            else:
                over_under = 'Under'
        else:
            over_under = 'Under'

        # Optional: filter by Over/Within/Under selection
        if selected_budget_status and over_under.lower() != selected_budget_status.lower():
            continue

        # Track counts after status filter applied
        if over_under == 'Over':
            over_count += 1
        elif over_under == 'Within':
            within_count += 1
        else:
            under_count += 1
        
        # Calculate summary totals
        total_budget_sum += budget
        total_spent_sum += spent
        if utilization > 90 and utilization <= 100:
            at_risk_count += 1
        # Cost breakdown by type
        cost_breakdown_map = defaultdict(float)
        for c in costs:
            cost_breakdown_map[c.get_cost_type_display()] += float(c.amount)
        cost_breakdown = [
            {'cost_type': k, 'total': v} for k, v in cost_breakdown_map.items()
        ]
        
        # Get assigned engineers
        assigned_engineers = p.assigned_engineers.all()
        engineer_names = [eng.get_full_name() or eng.username for eng in assigned_engineers]
        
        project_data.append({
            'id': p.id,
            'name': p.name,
            'prn': p.prn,
            'barangay': p.barangay,
            'project_cost': budget,
            'source_of_funds': p.source_of_funds,
            'status': p.status,
            'budget': budget,
            'spent': spent,
            'remaining': remaining,
            'utilization': utilization,
            'over_under': over_under,
            'cost_breakdown': cost_breakdown,
            'assigned_engineers': engineer_names,
            'engineer_count': len(engineer_names),
        })
        project_names.append(p.name)
        utilizations.append(utilization)
    
    # Chart data from period-filtered projects (projects_for_chart)
    chart_project_names = []
    chart_utilizations_list = []
    chart_over_count = 0
    chart_within_count = 0
    chart_under_count = 0
    for p in projects_for_chart:
        costs = ProjectCost.objects.filter(project=p)
        spent = sum([float(c.amount) for c in costs]) if costs else 0
        budget = float(p.project_cost) if p.project_cost else 0
        utilization = (spent / budget * 100) if budget > 0 else 0
        if budget > 0:
            if spent > budget:
                ou = 'Over'
            elif utilization >= 80:
                ou = 'Within'
            else:
                ou = 'Under'
        else:
            ou = 'Under'
        if ou == 'Over':
            chart_over_count += 1
        elif ou == 'Within':
            chart_within_count += 1
        else:
            chart_under_count += 1
        chart_project_names.append(p.name)
        chart_utilizations_list.append(utilization)
    chart_data = sorted(
        zip(chart_project_names, chart_utilizations_list, [float(getattr(p, 'project_cost') or 0) for p in projects_for_chart]),
        key=lambda x: x[1] if x[1] > 0 else x[2],
        reverse=True
    )[:20]
    chart_project_names = [item[0] for item in chart_data]
    chart_utilizations = [item[1] for item in chart_data]
    
    # Pagination - 20 items per page
    paginator = Paginator(project_data, 20)
    page_number = request.GET.get('page', 1)
    try:
        page_obj = paginator.get_page(page_number)
    except:
        page_obj = paginator.get_page(1)
    
    # Get paginated project data for display
    paginated_project_data = list(page_obj.object_list)
    
    # Get all unique barangays for filter dropdown
    all_barangays = Project.objects.values_list('barangay', flat=True).distinct().exclude(barangay__isnull=True).exclude(barangay='').order_by('barangay')
    
    # Full filtered project list for PDF export (PDFMake) - not paginated
    budget_pdf_projects = []
    for idx, p in enumerate(project_data):
        cost_breakdown_str = '; '.join([f"{c['cost_type']}: ₱{float(c['total']):,.2f}" for c in p.get('cost_breakdown', [])]) if p.get('cost_breakdown') else '—'
        budget_pdf_projects.append({
            'num': idx + 1,
            'prn': p.get('prn') or '—',
            'name': p.get('name') or '—',
            'barangay': p.get('barangay') or '—',
            'assigned_engineers': ', '.join(p.get('assigned_engineers', [])) if p.get('assigned_engineers') else '—',
            'budget': p.get('budget', 0),
            'spent': p.get('spent', 0),
            'utilization': p.get('utilization', 0),
            'over_under': p.get('over_under', '—'),
            'status': p.get('status', ''),
            'cost_breakdown': cost_breakdown_str,
        })
    budget_report_pdf_data = {
        'summary': {
            'total_budget': total_budget_sum,
            'total_spent': total_spent_sum,
            'total_remaining': total_budget_sum - total_spent_sum,
            'project_count': len(project_data),
            'over_count': over_count,
            'at_risk_count': at_risk_count,
            'within_count': within_count,
            'under_count': under_count,
        },
        'projects': budget_pdf_projects,
    }
    
    context = {
        'project_data': paginated_project_data,
        'project_data_json': json.dumps(paginated_project_data),  # For JavaScript modal
        'budget_report_pdf_data': budget_report_pdf_data,
        'project_names': json.dumps(chart_project_names),
        'utilizations': json.dumps(chart_utilizations),
        'over_count': over_count,
        'under_count': under_count,
        'within_count': within_count,
        'total_budget_sum': total_budget_sum,
        'total_spent_sum': total_spent_sum,
        'total_remaining_sum': total_budget_sum - total_spent_sum,
        'at_risk_count': at_risk_count,
        'page_obj': page_obj,  # For pagination controls
        'selected_barangay': selected_barangay,
        'selected_status': selected_status,
        'selected_start_date': selected_start_date,
        'selected_end_date': selected_end_date,
        'selected_cost_min': selected_cost_min,
        'selected_cost_max': selected_cost_max,
        'selected_budget_status': selected_budget_status,
        'selected_chart_period': selected_chart_period,
        'all_barangays': all_barangays,
        'chart_over_count': chart_over_count,
        'chart_within_count': chart_within_count,
        'chart_under_count': chart_under_count,
    }
    return render(request, 'monitoring/budget_reports.html', context)


@login_required
@prevent_project_engineer_access
def budget_reports_chart_data_api(request):
    """API: return chart data for Budget Reports (utilization or over/under) by period. No page reload."""
    from django.http import JsonResponse
    from projeng.models import Project, ProjectCost
    from django.db.models import Q
    from datetime import date, timedelta

    chart_type = (request.GET.get('chart') or '').strip().lower()
    period = (request.GET.get('period') or 'month').strip().lower()
    if chart_type not in ('utilization', 'overunder'):
        return JsonResponse({'error': 'Invalid chart'}, status=400)
    if period not in ('week', 'month', 'year'):
        period = 'month'

    if is_head_engineer(request.user) or is_finance_manager(request.user):
        projects = Project.objects.all()
    elif is_project_engineer(request.user):
        projects = Project.objects.filter(assigned_engineers=request.user)
    else:
        projects = Project.objects.none()

    selected_barangay = request.GET.get('barangay', '').strip()
    selected_status = request.GET.get('status', '').strip()
    selected_start_date = request.GET.get('start_date', '').strip()
    selected_end_date = request.GET.get('end_date', '').strip()
    selected_cost_min = request.GET.get('cost_min', '').strip()
    selected_cost_max = request.GET.get('cost_max', '').strip()
    selected_budget_status = request.GET.get('budget_status', '').strip()

    if selected_barangay:
        projects = projects.filter(barangay=selected_barangay)
    if selected_status:
        if selected_status == 'in_progress':
            projects = projects.filter(Q(status='in_progress') | Q(status='ongoing'))
        else:
            projects = projects.filter(status=selected_status)
    if selected_start_date:
        projects = projects.filter(start_date__gte=selected_start_date)
    if selected_end_date:
        projects = projects.filter(end_date__lte=selected_end_date)
    if selected_cost_min:
        try:
            projects = projects.filter(project_cost__gte=float(selected_cost_min))
        except (ValueError, TypeError):
            pass
    if selected_cost_max:
        try:
            projects = projects.filter(project_cost__lte=float(selected_cost_max))
        except (ValueError, TypeError):
            pass

    today = timezone.now().date() if timezone.is_naive(timezone.now()) else timezone.now().date()
    if period == 'week':
        period_start = today - timedelta(days=today.weekday())
        period_end = period_start + timedelta(days=6)
    elif period == 'year':
        period_start = date(today.year, 1, 1)
        period_end = date(today.year, 12, 31)
    else:
        period_start = date(today.year, today.month, 1)
        period_end = date(today.year, 12, 31) if today.month == 12 else (date(today.year, today.month + 1, 1) - timedelta(days=1))

    def in_period(p):
        start = getattr(p, 'start_date', None)
        end = getattr(p, 'end_date', None)
        if not start and not end:
            return True
        if start and end:
            return start <= period_end and end >= period_start
        if start:
            return start <= period_end
        if end:
            return end >= period_start
        return True

    projects_for_chart = [p for p in projects if in_period(p)]

    if chart_type == 'utilization':
        names = []
        utils = []
        budgets = []
        for p in projects_for_chart:
            costs = ProjectCost.objects.filter(project=p)
            spent = sum([float(c.amount) for c in costs]) if costs else 0
            budget = float(p.project_cost) if p.project_cost else 0
            util = (spent / budget * 100) if budget > 0 else 0
            names.append(p.name or '')
            utils.append(round(util, 2))
            budgets.append(budget)
        sorted_data = sorted(zip(names, utils, budgets), key=lambda x: x[1] if x[1] > 0 else x[2], reverse=True)[:20]
        return JsonResponse({
            'project_names': [x[0] for x in sorted_data],
            'utilizations': [x[1] for x in sorted_data],
        })
    else:
        over = within = under = 0
        for p in projects_for_chart:
            costs = ProjectCost.objects.filter(project=p)
            spent = sum([float(c.amount) for c in costs]) if costs else 0
            budget = float(p.project_cost) if p.project_cost else 0
            util = (spent / budget * 100) if budget > 0 else 0
            if budget > 0:
                if spent > budget:
                    ou = 'Over'
                elif util >= 80:
                    ou = 'Within'
                else:
                    ou = 'Under'
            else:
                ou = 'Under'
            if selected_budget_status and ou.lower() != selected_budget_status.lower():
                continue
            if ou == 'Over':
                over += 1
            elif ou == 'Within':
                within += 1
            else:
                under += 1
        return JsonResponse({'over': over, 'within': within, 'under': under})


@login_required
@head_engineer_required
def project_get_api(request, pk):
    """Get a single project's data as JSON for editing"""
    from django.http import JsonResponse
    from django.shortcuts import get_object_or_404
    from django.db.models import Sum
    import logging
    from decimal import Decimal
    
    logger = logging.getLogger(__name__)
    
    try:
        project = get_object_or_404(Project, pk=pk)
        
        # Safely get image URL
        image_url = ''
        if project.image:
            try:
                image_url = project.image.url
            except Exception as img_error:
                logger.warning(f"Error getting image URL for project {project.id}: {str(img_error)}")
                image_url = ''
        
        # Safely get assigned engineers
        assigned_engineers = []
        assigned_engineer_ids = []
        try:
            if hasattr(project, 'assigned_engineers'):
                assigned_engineers_list = project.assigned_engineers.all()
                assigned_engineers = [str(e) for e in assigned_engineers_list]
                assigned_engineer_ids = [e.id for e in assigned_engineers_list]
        except Exception as eng_error:
            logger.warning(f"Error getting assigned engineers for project {project.id}: {str(eng_error)}")
            assigned_engineers = []
            assigned_engineer_ids = []
        
        total_spent = ProjectCost.objects.filter(project=project).aggregate(total=Sum('amount')).get('total') or Decimal('0.00')
        remaining_funds = (project.project_cost - total_spent) if project.project_cost is not None else None

        project_data = {
            'id': project.id,
            'name': project.name or '',
            'prn': project.prn or '',
            'description': project.description or '',
            'barangay': project.barangay or '',
            'latitude': float(project.latitude) if project.latitude else '',
            'longitude': float(project.longitude) if project.longitude else '',
            'project_cost': f"{project.project_cost:.2f}" if project.project_cost is not None else '',
            'total_spent': f"{total_spent:.2f}",
            'remaining_funds': f"{remaining_funds:.2f}" if remaining_funds is not None else '',
            'source_of_funds': project.source_of_funds or '',
            'start_date': str(project.start_date) if project.start_date else '',
            'end_date': str(project.end_date) if project.end_date else '',
            'status': project.status or '',
            'image': image_url,
            'progress': getattr(project, 'progress', 0) or 0,
            'assigned_engineers': assigned_engineers,
            'assigned_engineer_ids': assigned_engineer_ids,
            'zone_type': project.zone_type or '',
        }
        
        return JsonResponse({'success': True, 'project': project_data})
    except Project.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'Project not found'}, status=404)
    except Exception as e:
        logger.error(f"Error getting project {pk}: {str(e)}", exc_info=True)
        return JsonResponse({'success': False, 'error': str(e)}, status=500)

def project_update_api(request, pk):
    return HttpResponse("project_update_api placeholder")

@login_required
@head_engineer_required
def project_delete_api(request, pk):
    """Delete a project and notify all relevant users"""
    from django.db import transaction
    from django.http import JsonResponse, HttpResponseNotAllowed
    from projeng.models import Project as ProjEngProject
    from monitoring.models import Project as MonitoringProject
    from projeng.utils import notify_head_engineers, notify_admins, notify_finance_managers
    from django.contrib.auth.models import User
    
    if request.method != 'POST' and request.method != 'DELETE':
        return HttpResponseNotAllowed(['POST', 'DELETE'])
    
    try:
        with transaction.atomic():
            project_name = None
            project_prn = None
            assigned_engineers = []
            
            # Try to find the project in either model
            projeng_project = ProjEngProject.objects.filter(pk=pk).first()
            monitoring_project = MonitoringProject.objects.filter(pk=pk).first()
            
            if projeng_project:
                # Get project details before deletion
                project_name = projeng_project.name
                project_prn = projeng_project.prn
                assigned_engineers = list(projeng_project.assigned_engineers.all())
                
                # Delete the project (this will cascade to related objects)
                projeng_project.delete()
                
                # Also delete from monitoring if it exists
                if project_prn:
                    MonitoringProject.objects.filter(prn=project_prn).delete()
                else:
                    MonitoringProject.objects.filter(name=project_name).delete()
                    
            elif monitoring_project:
                # Get project details before deletion
                project_name = monitoring_project.name
                project_prn = monitoring_project.prn
                
                # Try to find corresponding projeng project to get assigned engineers
                if project_prn:
                    projeng_project = ProjEngProject.objects.filter(prn=project_prn).first()
                    if projeng_project:
                        assigned_engineers = list(projeng_project.assigned_engineers.all())
                        projeng_project.delete()
                
                # Delete from monitoring
                monitoring_project.delete()
            else:
                return JsonResponse({'success': False, 'error': 'Project not found'}, status=404)
            
            # Notify all relevant users about the deletion
            deleter_name = request.user.get_full_name() or request.user.username
            project_display = f"{project_name}" + (f" (PRN: {project_prn})" if project_prn else "")
            
            # Notify Head Engineers and Admins
            message = f"Project '{project_display}' has been deleted by {deleter_name}"
            notify_head_engineers(message)
            notify_admins(message)
            
            # Notify Finance Managers
            notify_finance_managers(message)
            
            # Notify assigned Project Engineers (with duplicate check)
            from projeng.models import Notification
            from django.utils import timezone
            from datetime import timedelta
            
            engineer_message = f"Project '{project_display}' that you were assigned to has been deleted by {deleter_name}"
            recent_time = timezone.now() - timedelta(seconds=10)
            
            from django.db.models import Q
            for engineer in assigned_engineers:
                # Check for duplicates before creating notification
                engineer_duplicate = Notification.objects.filter(
                    Q(recipient=engineer) &
                    Q(message__icontains=project_display) &
                    Q(message__icontains="that you were assigned to has been deleted") &
                    Q(created_at__gte=recent_time)
                ).exists()
                
                if not engineer_duplicate:
                    Notification.objects.create(
                        recipient=engineer,
                        message=engineer_message
                    )
            
            # Phase 3: Also broadcast via WebSocket (parallel to SSE)
            try:
                from projeng.channels_utils import broadcast_project_deleted
                broadcast_project_deleted(project_name, project_prn)
            except Exception as e:
                print(f"⚠️  WebSocket broadcast failed (SSE still works): {e}")
            
            return JsonResponse({
                'success': True,
                'message': f'Project "{project_name}" deleted successfully'
            })
            
    except Exception as e:
        import traceback
        traceback.print_exc()
        return JsonResponse({
            'success': False,
            'error': f'An error occurred while deleting the project: {str(e)}'
        }, status=500)

@login_required
@prevent_project_engineer_access
def delayed_projects(request):
    """View for displaying delayed projects with filters"""
    from projeng.models import Project
    from django.core.paginator import Paginator
    from django.db.models import Q
    
    # Get delayed projects
    if is_head_engineer(request.user) or is_finance_manager(request.user):
        projects = Project.objects.filter(status='delayed')
    elif is_project_engineer(request.user):
        projects = Project.objects.filter(status='delayed', assigned_engineers=request.user)
    else:
        projects = Project.objects.none()
    
    # Apply filters
    barangay_filter = request.GET.get('barangay')
    search_query = request.GET.get('search', '').strip()
    sort_by = request.GET.get('sort', 'created_at')
    
    # Barangay filter
    if barangay_filter:
        projects = projects.filter(barangay=barangay_filter)
    
    # Search filter
    if search_query:
        projects = projects.filter(
            Q(name__icontains=search_query) |
            Q(prn__icontains=search_query) |
            Q(description__icontains=search_query) |
            Q(barangay__icontains=search_query)
        )
    
    # Sort
    if sort_by == 'name':
        projects = projects.order_by('name')
    elif sort_by == 'barangay':
        projects = projects.order_by('barangay')
    else:  # default to created_at
        projects = projects.order_by('-created_at')
    
    total_delayed = projects.count()
    
    # Pagination
    paginator = Paginator(projects, 20)  # Show 20 projects per page
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    return render(request, 'monitoring/delayed_projects.html', {
        'page_obj': page_obj,
        'total_delayed': total_delayed,
    })

def project_engineer_analytics(request, pk):
    return HttpResponse("project_engineer_analytics placeholder")

@login_required
@head_engineer_required
def head_engineer_analytics(request):
    from projeng.models import Project, ProjectProgress, SourceOfFunds, ProjectType
    from django.db.models import Max, Q
    from django.core.paginator import Paginator
    from django.utils import timezone
    from decimal import Decimal, InvalidOperation
    import json

    # Base queryset (will be narrowed by filters)
    base_projects = Project.objects.all().select_related('project_type')

    # Seed SourceOfFunds master list if empty (so dropdown isn't blank)
    if not SourceOfFunds.objects.exists():
        existing_sources = (
            Project.objects.exclude(source_of_funds__isnull=True)
            .exclude(source_of_funds='')
            .values_list('source_of_funds', flat=True)
            .distinct()
        )
        for name in existing_sources:
            cleaned = (name or '').strip()
            if cleaned and not SourceOfFunds.objects.filter(name__iexact=cleaned).exists():
                SourceOfFunds.objects.create(name=cleaned)

    source_of_funds_options = SourceOfFunds.objects.filter(is_active=True).order_by('name')
    project_type_options = ProjectType.objects.all().order_by('name')

    # Read filters (match Head Engineer filter bar)
    barangay_filter = (request.GET.get('barangay') or '').strip()
    status_filter = (request.GET.get('status') or '').strip()
    duration_filter = (request.GET.get('duration') or '').strip()
    source_of_funds_filter = (request.GET.get('source_of_funds') or '').strip()
    project_type_filter = (request.GET.get('project_type') or '').strip()
    cost_min_raw = (request.GET.get('cost_min') or '').strip()
    cost_max_raw = (request.GET.get('cost_max') or '').strip()
    search_query = (request.GET.get('search') or '').strip()

    # Apply NON-status filters first (so summary cards reflect this scope)
    if barangay_filter:
        base_projects = base_projects.filter(barangay=barangay_filter)

    if source_of_funds_filter:
        base_projects = base_projects.filter(source_of_funds__iexact=source_of_funds_filter)

    if project_type_filter:
        try:
            base_projects = base_projects.filter(project_type_id=int(project_type_filter))
        except (TypeError, ValueError):
            project_type_filter = ''

    if cost_min_raw:
        try:
            base_projects = base_projects.filter(project_cost__gte=Decimal(cost_min_raw))
        except (InvalidOperation, ValueError):
            cost_min_raw = ''
    if cost_max_raw:
        try:
            base_projects = base_projects.filter(project_cost__lte=Decimal(cost_max_raw))
        except (InvalidOperation, ValueError):
            cost_max_raw = ''

    if search_query:
        base_projects = base_projects.filter(
            Q(name__icontains=search_query) |
            Q(prn__icontains=search_query) |
            Q(description__icontains=search_query) |
            Q(barangay__icontains=search_query)
        )

    # Duration filter (based on project duration start_date→end_date)
    if duration_filter:
        projects_with_dates = base_projects.filter(start_date__isnull=False, end_date__isnull=False)
        matching_ids = []
        for proj in projects_with_dates.only('id', 'start_date', 'end_date'):
            if not proj.start_date or not proj.end_date:
                continue
            duration_days = (proj.end_date - proj.start_date).days
            if duration_filter == 'lt6' and duration_days < 180:
                matching_ids.append(proj.id)
            elif duration_filter == '6to12' and 180 <= duration_days <= 365:
                matching_ids.append(proj.id)
            elif duration_filter == 'gt12' and duration_days > 365:
                matching_ids.append(proj.id)
        base_projects = base_projects.filter(id__in=matching_ids) if matching_ids else base_projects.none()

    # Build latest progress map for the current filtered scope (for cards + status filtering)
    filtered_ids = list(base_projects.values_list('id', flat=True))
    latest_progress_all = {}
    if filtered_ids:
        latest_progress_qs = ProjectProgress.objects.filter(project_id__in=filtered_ids).values('project_id').annotate(
            latest_date=Max('date'),
            latest_created=Max('created_at'),
        )
        for item in latest_progress_qs:
            latest = ProjectProgress.objects.filter(
                project_id=item['project_id'],
                date=item['latest_date'],
                created_at=item['latest_created']
            ).order_by('-created_at').first()
            if latest and latest.percentage_complete is not None:
                latest_progress_all[item['project_id']] = int(latest.percentage_complete)

    today = timezone.now().date()
    total_projects_all = base_projects.count()
    completed_projects = 0
    ongoing_projects = 0
    planned_projects = 0
    delayed_projects = 0

    calculated_statuses = {}
    for p in list(base_projects):
        progress = latest_progress_all.get(p.id, 0)
        stored_status = p.status or ''

        calculated_status = stored_status
        if progress >= 99:
            calculated_status = 'completed'
            completed_projects += 1
        elif stored_status == 'delayed':
            calculated_status = 'delayed'
            delayed_projects += 1
        elif progress < 99 and p.end_date and p.end_date < today and stored_status in ['in_progress', 'ongoing']:
            calculated_status = 'delayed'
            delayed_projects += 1
        elif stored_status in ['in_progress', 'ongoing']:
            calculated_status = 'in_progress'
            ongoing_projects += 1
        elif stored_status in ['planned', 'pending']:
            calculated_status = 'planned'
            planned_projects += 1
        elif stored_status == 'completed':
            calculated_status = 'completed'
            completed_projects += 1

        calculated_statuses[p.id] = calculated_status
    
    # Apply status filter on top of the filtered base scope
    projects = base_projects
    if status_filter:
        matching_ids = [pid for pid, st in calculated_statuses.items() if st == status_filter]
        projects = projects.filter(id__in=matching_ids) if matching_ids else projects.none()
    
    # Order by created_at descending
    projects = projects.order_by('-created_at')
    
    # Get all project IDs for batch queries (after filtering)
    all_project_ids_for_page = list(projects.values_list('id', flat=True))
    
    # Batch fetch latest progress for filtered projects
    latest_progress_dict = {}
    if all_project_ids_for_page:
        # Get the latest progress update for each project
        for project_id in all_project_ids_for_page:
            latest = ProjectProgress.objects.filter(
                project_id=project_id
            ).order_by('-date', '-created_at').first()
            if latest:
                latest_progress_dict[project_id] = latest
    
    # Batch fetch assigned engineers for all projects
    from django.db.models import Prefetch
    projects_with_engineers = Project.objects.filter(
        id__in=all_project_ids_for_page
    ).prefetch_related('assigned_engineers')
    engineers_dict = {}
    for proj in projects_with_engineers:
        engineers_dict[proj.id] = [eng.username for eng in proj.assigned_engineers.all()]
    
    # Prepare project list for the table and JS (calculate status for all projects first)
    projects_list = []
    for p in projects:
        # Get latest progress from batch query
        latest_progress = latest_progress_dict.get(p.id)
        
        # Calculate progress - priority: 1) completed status = 100%, 2) latest progress update, 3) project.progress field, 4) 0
        if p.status == 'completed':
            progress = 100
        elif latest_progress and latest_progress.percentage_complete is not None:
            # Ensure progress is between 0 and 100
            try:
                progress = max(0, min(100, int(latest_progress.percentage_complete)))
            except (ValueError, TypeError):
                progress = 0
        elif hasattr(p, 'progress') and p.progress is not None:
            # Fallback to project's direct progress field
            try:
                progress = max(0, min(100, int(p.progress)))
            except (ValueError, TypeError):
                progress = 0
        else:
            progress = 0
        
        # Calculate actual status dynamically (including delayed)
        stored_status = p.status or ''
        calculated_status = stored_status
        
        # Priority: completed > delayed > in_progress > planned
        if progress >= 99:
            calculated_status = 'completed'
        elif stored_status == 'delayed':
            calculated_status = 'delayed'
        elif progress < 99 and p.end_date and p.end_date < today and stored_status in ['in_progress', 'ongoing']:
            # Project is delayed if: end_date passed, progress < 99%, and status is in_progress/ongoing
            calculated_status = 'delayed'
        elif stored_status in ['in_progress', 'ongoing']:
            calculated_status = 'in_progress'
        elif stored_status in ['planned', 'pending']:
            calculated_status = 'planned'
        
        # Status display
        status_display = (
            'Completed' if calculated_status == 'completed' else
            'Delayed' if calculated_status == 'delayed' else
            'Ongoing' if calculated_status in ['in_progress', 'ongoing'] else
            'Planned' if calculated_status in ['planned', 'pending'] else
            calculated_status.title()
        )
        
        # Assigned engineers from batch query
        assigned_to = engineers_dict.get(p.id, [])
        
        projects_list.append({
            'id': p.id,
            'name': p.name,
            'prn': p.prn or '',
            'barangay': p.barangay or '',
            'project_type': p.project_type.name if (getattr(p, 'project_type', None) and p.project_type) else '',
            'source_of_funds': p.source_of_funds or '',
            'total_progress': progress,
            'status': calculated_status,  # Use calculated_status instead of stored status
            'status_display': status_display,
            'start_date': str(p.start_date) if p.start_date else '',
            'end_date': str(p.end_date) if p.end_date else '',
            'assigned_to': assigned_to,
        })
    
    # If filtering by delayed status, filter the list to only show delayed projects
    if status_filter == 'delayed':
        projects_list = [p for p in projects_list if p['status'] == 'delayed']
    
    # Pagination - 15 items per page (after filtering if needed)
    from django.core.paginator import Paginator
    paginator = Paginator(projects_list, 15)
    page_number = request.GET.get('page', 1)
    page_obj = paginator.get_page(page_number)
    
    # Update projects_list to only include current page items
    projects_list = list(page_obj.object_list)
    
    context = {
        'projects': projects_list,  # for Django template rendering
        'projects_json': json.dumps(projects_list),  # for JS
        'page_obj': page_obj,  # for pagination
        'total_projects': total_projects_all,
        'completed_projects': completed_projects,
        'ongoing_projects': ongoing_projects,
        'planned_projects': planned_projects,
        'delayed_projects': delayed_projects,
        'status_filter': status_filter,
        'barangay_filter': barangay_filter,
        'search_query': search_query,
        'user_role': 'head_engineer',
        'duration_filter': duration_filter,
        'source_of_funds_filter': source_of_funds_filter,
        'source_of_funds_options': source_of_funds_options,
        'project_type_options': project_type_options,
        'project_type_filter': project_type_filter,
        'cost_min': cost_min_raw,
        'cost_max': cost_max_raw,
    }
    return render(request, 'monitoring/analytics.html', context)

def project_detail(request, pk):
    return HttpResponse("project_detail placeholder")


def _file_to_data_url(file_field):
    """Convert a FileField/ImageField to base64 JPEG dataURL for pdfMake.
    Uses Pillow to normalize to JPEG (pdfMake supports JPEG/PNG; WebP/etc cause 'Unknown image format').
    """
    import base64
    import io
    if not file_field or not getattr(file_field, 'name', None):
        return None
    try:
        from PIL import Image
        with file_field.open('rb') as fh:
            content = fh.read()
        if not content:
            return None
        img = Image.open(io.BytesIO(content))
        if img.mode in ('RGBA', 'P'):
            img = img.convert('RGB')
        elif img.mode != 'RGB':
            img = img.convert('RGB')
        buf = io.BytesIO()
        img.save(buf, format='JPEG', quality=85, optimize=True)
        b64 = base64.b64encode(buf.getvalue()).decode('ascii')
        return f'data:image/jpeg;base64,{b64}'
    except Exception:
        return None


def _build_uploaded_images_for_pdf_urls(request, progress_updates, image_documents, project):
    """Build image list with absolute URLs for server-side xhtml2pdf (fetches URLs; data URIs not well supported)."""
    images = []
    for u in progress_updates:
        for photo in (u.photos.all() if hasattr(u, 'photos') else []):
            try:
                url = photo.image.url if photo.image else None
                if url:
                    images.append({
                        'url': request.build_absolute_uri(url),
                        'caption': u.date.strftime('%Y-%m-%d') if getattr(u, 'date', None) else 'Progress',
                    })
            except (ValueError, OSError):
                pass
    for doc in image_documents or []:
        try:
            url = doc.file.url if doc.file else None
            if url:
                images.append({
                    'url': request.build_absolute_uri(url),
                    'caption': (doc.name or (doc.file.name if doc.file else '') or 'Document')[:30],
                })
        except (ValueError, OSError):
            pass
    if project and getattr(project, 'image', None) and project.image:
        try:
            url = project.image.url
            if url:
                images.append({
                    'url': request.build_absolute_uri(url),
                    'caption': 'Project',
                })
        except (ValueError, OSError):
            pass
    return images


def _build_uploaded_images_list(request, progress_updates, image_documents, project):
    """Collect all uploaded images as base64 dataURLs for PDF export (pdfMake requires dataURL, not remote URLs)."""
    images = []
    # Progress photos from updates
    for u in progress_updates:
        for photo in (u.photos.all() if hasattr(u, 'photos') else []):
            try:
                data_url = _file_to_data_url(photo.image)
                if data_url:
                    images.append({
                        'url': data_url,
                        'caption': u.date.strftime('%Y-%m-%d') if getattr(u, 'date', None) else 'Progress',
                    })
            except (ValueError, OSError):
                pass
    # Image documents
    for doc in image_documents or []:
        try:
            data_url = _file_to_data_url(doc.file)
            if data_url:
                images.append({
                    'url': data_url,
                    'caption': (doc.name or (doc.file.name if doc.file else '') or 'Document')[:30],
                })
        except (ValueError, OSError):
            pass
    # Project main image
    if project and getattr(project, 'image', None) and project.image:
        try:
            data_url = _file_to_data_url(project.image)
            if data_url:
                images.append({
                    'url': data_url,
                    'caption': 'Project',
                })
        except (ValueError, OSError):
            pass
    return images


@login_required
@head_engineer_required
def head_engineer_project_detail(request, pk):
    import logging
    import json
    from projeng.models import Project, ProjectProgress, ProjectCost, Notification, ProjectDocument
    from django.contrib.auth.models import User
    from projeng.utils import get_project_from_notification
    from django.http import HttpResponse
    
    logger = logging.getLogger(__name__)
    
    try:
        project = Project.objects.get(pk=pk)
    except Project.DoesNotExist:
        return HttpResponse('Project not found.', status=404)
    except Exception as e:
        logger.error(f"Error fetching project {pk}: {e}")
        return HttpResponse(f'Error loading project: {str(e)}', status=500)
    
    try:
        # Auto-mark notifications as read for this project
        # Get all unread notifications for the user
        unread_notifications = Notification.objects.filter(
            recipient=request.user,
            is_read=False
        )
        
        # Mark notifications related to this project as read
        for notification in unread_notifications:
            try:
                if notification.message:
                    project_id = get_project_from_notification(notification.message)
                    if project_id == project.id:
                        notification.is_read = True
                        notification.save()
            except Exception as e:
                # Log error but continue processing
                logger.warning(f"Error processing notification {notification.id}: {e}")
                continue
        
        # Get all progress updates - order by date and id to avoid duplicates and ensure consistent ordering
        progress_updates = ProjectProgress.objects.filter(project=project).prefetch_related('photos').order_by('date', 'id').distinct()
        assigned_to = list(project.assigned_engineers.values_list('username', flat=True))
        # Get all cost entries with created_by prefetched
        costs = ProjectCost.objects.filter(project=project).select_related('created_by').order_by('date')
        # Get all documents with uploaded_by prefetched
        all_documents = ProjectDocument.objects.filter(project=project).select_related('uploaded_by').order_by('-uploaded_at')
        
        # Separate images from other documents
        image_extensions = ['.jpg', '.jpeg', '.png', '.gif', '.webp', '.bmp', '.svg']
        image_documents = []
        other_documents = []
        
        for doc in all_documents:
            file_name_lower = doc.file.name.lower()
            is_image = any(file_name_lower.endswith(ext) for ext in image_extensions)
            if is_image:
                image_documents.append(doc)
            else:
                other_documents.append(doc)
        
        documents = all_documents  # Keep for backward compatibility
        # Analytics & summary
        latest_progress = progress_updates.last() if progress_updates else None
        
        # Calculate total cost safely
        total_cost = 0
        try:
            if costs:
                for c in costs:
                    try:
                        total_cost += float(c.amount) if c.amount else 0
                    except (ValueError, TypeError):
                        continue
        except Exception as e:
            logger.warning(f"Error calculating total cost: {e}")
            total_cost = 0
        
        # Calculate budget utilization safely
        try:
            if project.project_cost:
                project_cost_float = float(project.project_cost)
                if project_cost_float > 0:
                    budget_utilization = (total_cost / project_cost_float) * 100
                else:
                    budget_utilization = 0
            else:
                budget_utilization = 0
        except (ValueError, TypeError) as e:
            logger.warning(f"Error calculating budget utilization: {e}")
            budget_utilization = 0
        
        timeline_data = {
            'start_date': project.start_date,
            'end_date': project.end_date,
            'days_elapsed': (project.end_date - project.start_date).days if project.start_date and project.end_date else None,
            'total_days': (project.end_date - project.start_date).days if project.start_date and project.end_date else None,
        }
        
        # Serializable report data for client-side PDFMake (comprehensive layout)
        from django.utils import timezone as tz
        from collections import defaultdict
        today = tz.now().date()
        total_days = (project.end_date - project.start_date).days if project.start_date and project.end_date else 0
        days_elapsed = (today - project.start_date).days if project.start_date else 0
        days_remaining = total_days - days_elapsed if total_days > 0 else 0
        budget = float(project.project_cost) if project.project_cost else 0
        remaining_budget = budget - total_cost if budget else 0
        expected_progress = min(100.0, (days_elapsed / total_days * 100.0)) if total_days > 0 else None
        progress_variance = (latest_progress.percentage_complete - expected_progress) if latest_progress and expected_progress is not None else None
        performance_ratio = (latest_progress.percentage_complete / expected_progress) if latest_progress and expected_progress and expected_progress > 0 else None
        performance_label = 'N/A'
        if performance_ratio is not None:
            if performance_ratio >= 1.05: performance_label = 'Ahead of schedule'
            elif performance_ratio >= 0.95: performance_label = 'On schedule'
            else: performance_label = 'Behind schedule'
        budget_status_label = 'UNDER BUDGET'
        if budget_utilization > 100: budget_status_label = 'OVER BUDGET'
        elif budget_utilization >= 90: budget_status_label = 'AT RISK'
        cost_breakdown = defaultdict(float)
        for c in costs:
            cost_breakdown[c.get_cost_type_display()] += float(c.amount or 0)
        budget_requests_list = []
        try:
            from projeng.models import BudgetRequest
            for br in BudgetRequest.objects.filter(project=project).order_by('-created_at')[:10]:
                budget_requests_list.append({
                    'date': br.created_at.strftime('%Y-%m-%d'),
                    'requested': float(br.requested_amount or 0),
                    'status': br.get_status_display(),
                    'approved': float(br.approved_amount or 0) if br.approved_amount else None,
                    'by': br.requested_by.get_full_name() or br.requested_by.username if br.requested_by else '',
                    'reason': (br.reason or '')[:100],
                })
        except Exception:
            pass
        report_data = {
            'project': {
                'name': project.name or '',
                'prn': project.prn or '',
                'barangay': project.barangay or '',
                'status': project.get_status_display() if hasattr(project, 'get_status_display') else (project.status or ''),
                'start_date': project.start_date.strftime('%Y-%m-%d') if project.start_date else '',
                'end_date': project.end_date.strftime('%Y-%m-%d') if project.end_date else '',
                'project_cost': budget,
                'source_of_funds': project.source_of_funds or '',
                'description': (project.description or '')[:200],
            },
            'assigned_engineers': [u.get_full_name() or u.username for u in project.assigned_engineers.all()],
            'latest_progress_pct': latest_progress.percentage_complete if latest_progress else 0,
            'total_cost': float(total_cost),
            'budget': budget,
            'remaining_budget': remaining_budget,
            'budget_utilization': float(budget_utilization),
            'budget_status_label': budget_status_label,
            'days_elapsed': days_elapsed,
            'total_days': total_days,
            'days_remaining': days_remaining,
            'expected_progress': expected_progress,
            'progress_variance': progress_variance,
            'performance_label': performance_label,
            'cost_breakdown': dict(cost_breakdown),
            'budget_requests': budget_requests_list,
            'generated_by': request.user.get_full_name() or request.user.username,
            'generated_at': tz.now().strftime('%B %d, %Y at %I:%M %p'),
            'progress_updates': [
                {
                    'date': u.date.strftime('%Y-%m-%d') if getattr(u, 'date', None) else '',
                    'time': u.created_at.strftime('%I:%M %p') if getattr(u, 'created_at', None) else '',
                    'percentage_complete': getattr(u, 'percentage_complete', 0),
                    'description': (getattr(u, 'description', None) or '')[:80],
                    'engineer': u.created_by.get_full_name() or getattr(u.created_by, 'username', '') if getattr(u, 'created_by', None) else '',
                    'photo_count': u.photos.count() if hasattr(u, 'photos') and hasattr(u.photos, 'count') else 0,
                }
                for u in progress_updates
            ],
            'costs': [
                {
                    'date': c.date.strftime('%Y-%m-%d') if getattr(c, 'date', None) else '',
                    'time': c.created_at.strftime('%I:%M %p') if getattr(c, 'created_at', None) else '',
                    'cost_type': c.get_cost_type_display() if hasattr(c, 'get_cost_type_display') else '',
                    'description': (getattr(c, 'description', None) or '')[:60],
                    'amount': float(getattr(c, 'amount', 0) or 0),
                    'engineer': c.created_by.get_full_name() or getattr(c.created_by, 'username', '') if getattr(c, 'created_by', None) else '',
                }
                for c in costs
            ],
            'uploaded_images': _build_uploaded_images_list(request, progress_updates, image_documents, project),
        }
        
        context = {
            'project': project,
            'projeng_project': project,  # Pass project as projeng_project for template compatibility
            'progress_updates': progress_updates,
            'assigned_to': assigned_to,
            'costs': costs,
            'documents': documents,  # All documents for backward compatibility
            'image_documents': image_documents,  # Images only
            'other_documents': other_documents,  # Non-image documents
            'latest_progress': latest_progress,
            'total_cost': total_cost,
            'budget_utilization': budget_utilization,
            'timeline_data': timeline_data,
            'report_data': report_data,
        }
        
        return render(request, 'monitoring/head_engineer_project_detail.html', context)
    
    except Exception as e:
        logger.error(f"Error in head_engineer_project_detail for project {pk}: {e}", exc_info=True)
        return HttpResponse(f'Server error: {str(e)}', status=500)

@login_required
@head_engineer_required
def head_dashboard_card_data_api(request):
    return HttpResponse("head_dashboard_card_data_api placeholder")

def barangay_geojson_view(request):
    geojson_path = os.path.join(settings.BASE_DIR, 'static', 'data', 'tagum_barangays.geojson')
    with open(geojson_path, 'r', encoding='utf-8') as f:
        geojson_data = json.load(f)
    return JsonResponse(geojson_data, safe=False)


def tagum_city_boundary_geojson_view(request):
    """Serve whole Tagum City boundary (OSM admin boundary) - single outline, no barangays."""
    geojson_path = os.path.join(settings.BASE_DIR, 'wholetagumexport.geojson')
    if not os.path.isfile(geojson_path):
        geojson_path = os.path.join(settings.BASE_DIR, 'static', 'data', 'tagum_city_boundary.geojson')
    with open(geojson_path, 'r', encoding='utf-8') as f:
        geojson_data = json.load(f)
    return JsonResponse(geojson_data, safe=False)

def export_project_timeline_pdf(request, pk):
    """Export project timeline as PDF"""
    from projeng.models import Project as ProjEngProject, ProjectProgress as ProjEngProjectProgress
    try:
        # Try to get project from projeng first (since head_engineer_project_detail uses projeng Project)
        try:
            project = ProjEngProject.objects.get(pk=pk)
            projeng_project = project
        except ProjEngProject.DoesNotExist:
            # Fallback to monitoring project
            project = Project.objects.get(pk=pk)
            projeng_project = None
            if project.prn:
                try:
                    projeng_project = ProjEngProject.objects.get(prn=project.prn)
                except ProjEngProject.DoesNotExist:
                    pass
    except Project.DoesNotExist:
        return HttpResponse('Project not found.', status=404)
    
    # Get all progress updates - order by date and id to avoid duplicates
    if projeng_project:
        progress_updates = ProjEngProjectProgress.objects.filter(project=projeng_project).order_by('date', 'id').distinct()
    else:
        progress_updates = []
    
    # Get costs - try from projeng project first, then monitoring project
    if projeng_project:
        costs = ProjectCost.objects.filter(project=projeng_project).order_by('date')
    else:
        costs = ProjectCost.objects.filter(project=project).order_by('date') if hasattr(project, 'projectcost_set') else []

    # Compute consistent summary metrics for clearer report presentation (panel #27/#29)
    total_cost = sum([float(c.amount) for c in costs]) if costs else 0.0
    budget = float(getattr(project, 'project_cost', 0) or 0)
    remaining_budget = (budget - total_cost) if budget > 0 else 0.0
    budget_utilization = (total_cost / budget * 100) if budget > 0 else 0.0
    budget_used_pct = float(budget_utilization) if budget_utilization else 0.0

    latest_progress = progress_updates.last() if progress_updates else None
    total_progress = latest_progress.percentage_complete if latest_progress else 0

    today = timezone.now().date()
    total_days = (project.end_date - project.start_date).days if getattr(project, 'start_date', None) and getattr(project, 'end_date', None) else 0
    days_elapsed = (today - project.start_date).days if getattr(project, 'start_date', None) else 0
    days_remaining = (total_days - days_elapsed) if total_days > 0 else 0

    expected_progress = None
    progress_variance = None
    performance_ratio = None
    if getattr(project, 'start_date', None) and getattr(project, 'end_date', None) and total_days > 0:
        elapsed_days = max(0, min(total_days, days_elapsed))
        expected_progress = min(100.0, (elapsed_days / total_days) * 100.0)
        progress_variance = float(total_progress) - float(expected_progress)
        performance_ratio = (float(total_progress) / float(expected_progress)) if expected_progress > 0 else None

    efficiency_ratio = (float(total_progress) / budget_used_pct) if budget_used_pct > 0 else None

    budget_status_label = 'UNDER BUDGET'
    if budget_used_pct > 100:
        budget_status_label = 'OVER BUDGET'
    elif budget_used_pct >= 90:
        budget_status_label = 'AT RISK'

    # Achieve Satisfaction (0-100), same basis as comprehensive report
    satisfaction_score = None
    satisfaction_label = 'N/A'
    try:
        budget_component = None
        if efficiency_ratio is not None:
            budget_component = max(0.0, 100.0 - min(100.0, abs(1.0 - float(efficiency_ratio)) * 100.0))
        schedule_component = None
        if progress_variance is not None:
            schedule_component = max(0.0, 100.0 - abs(float(progress_variance)))
        if budget_component is not None and schedule_component is not None:
            base = (0.5 * schedule_component) + (0.5 * budget_component)
        elif budget_component is not None:
            base = budget_component
        else:
            base = None
        if base is not None and budget_used_pct > 100:
            base = base - 10.0
        if base is not None:
            satisfaction_score = round(max(0.0, min(100.0, float(base))), 1)
            if satisfaction_score >= 85:
                satisfaction_label = 'High'
            elif satisfaction_score >= 70:
                satisfaction_label = 'Moderate'
            else:
                satisfaction_label = 'Low'
    except Exception:
        satisfaction_score = None
        satisfaction_label = 'N/A'
    
    # If xhtml2pdf is unavailable, return a friendly message
    if pisa is None:
        return HttpResponse('PDF export is temporarily unavailable (missing xhtml2pdf/reportlab).', content_type='text/plain')
    
    # Render the HTML template for the PDF
    template_path = 'monitoring/project_timeline_pdf.html'
    template = get_template(template_path)
    context = {
        'project': project,
        'projeng_project': projeng_project,
        'progress_updates': progress_updates,
        'costs': costs,
        'total_cost': total_cost,
        'budget': budget,
        'remaining_budget': remaining_budget,
        'budget_utilization': budget_utilization,
        'budget_status_label': budget_status_label,
        'total_progress': total_progress,
        'expected_progress': expected_progress,
        'progress_variance': progress_variance,
        'performance_ratio': performance_ratio,
        'efficiency_ratio': efficiency_ratio,
        'satisfaction_score': satisfaction_score,
        'satisfaction_label': satisfaction_label,
        'days_elapsed': days_elapsed,
        'total_days': total_days,
        'days_remaining': days_remaining,
    }
    html = template.render(context)
    
    # Create a PDF
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="project_{project.id}_timeline_{timezone.now().strftime("%Y%m%d")}.pdf"'
    
    # Create a file-like object to write the PDF data to
    buffer = io.BytesIO()
    
    # Create the PDF object, and write the HTML to it
    pisa_status = pisa.CreatePDF(html, dest=buffer)
    
    # If there were no errors, return the PDF file
    if not pisa_status.err:
        return HttpResponse(buffer.getvalue(), content_type='application/pdf')
    
    # If there were errors, return an error message
    return HttpResponse('We had some errors generating the PDF.', content_type='text/plain')

@login_required
def export_reports_csv(request):
    """Export monitoring reports as CSV"""
    from django.db.models import Q
    
    # Role-based queryset
    if is_head_engineer(request.user) or is_finance_manager(request.user):
        projects = Project.objects.all()
    elif is_project_engineer(request.user):
        projects = Project.objects.filter(assigned_engineers=request.user)
    else:
        projects = Project.objects.none()
    
    # Apply filters from request parameters
    barangay_filter = request.GET.get('barangay', '').strip()
    status_filter = request.GET.get('status', '').strip()
    start_date_filter = request.GET.get('start_date', '').strip()
    end_date_filter = request.GET.get('end_date', '').strip()
    
    # Filter by barangay
    if barangay_filter:
        projects = projects.filter(barangay=barangay_filter)
    
    # Filter by status
    if status_filter:
        # Handle "in_progress" to match both "in_progress" and "ongoing" statuses
        if status_filter == 'in_progress':
            projects = projects.filter(Q(status='in_progress') | Q(status='ongoing'))
        else:
            projects = projects.filter(status=status_filter)
    
    # Filter by start date
    if start_date_filter:
        try:
            from datetime import datetime
            start_date = datetime.strptime(start_date_filter, '%Y-%m-%d').date()
            projects = projects.filter(start_date__gte=start_date)
        except (ValueError, TypeError):
            pass  # Invalid date format, ignore filter
    
    # Filter by end date
    if end_date_filter:
        try:
            from datetime import datetime
            end_date = datetime.strptime(end_date_filter, '%Y-%m-%d').date()
            projects = projects.filter(end_date__lte=end_date)
        except (ValueError, TypeError):
            pass  # Invalid date format, ignore filter
    
    # Filter by source of funds
    source_of_funds_filter = request.GET.get('source_of_funds', '').strip()
    if source_of_funds_filter:
        projects = projects.filter(source_of_funds__iexact=source_of_funds_filter)
    
    # Filter by project cost range
    cost_min_filter = request.GET.get('cost_min', '').strip()
    cost_max_filter = request.GET.get('cost_max', '').strip()
    if cost_min_filter:
        try:
            cost_min = float(cost_min_filter)
            projects = projects.filter(project_cost__gte=cost_min)
        except (ValueError, TypeError):
            pass
    if cost_max_filter:
        try:
            cost_max = float(cost_max_filter)
            projects = projects.filter(project_cost__lte=cost_max)
        except (ValueError, TypeError):
            pass
    
    # Create the HttpResponse object with the appropriate CSV header
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = f'attachment; filename="projects_report_{timezone.now().strftime("%Y%m%d")}.csv"'
    
    writer = csv.writer(response)
    # Write the header row
    writer.writerow(['#', 'PRN#', 'Name of Project', 'Project Description', 'Barangay', 'Project Cost', 'Source of Funds', 'Date Started', 'Date Ended', 'Status'])
    
    # Write data rows
    for i, project in enumerate(projects):
        writer.writerow([
            i + 1,
            project.prn or '',
            project.name or '',
            project.description or '',
            project.barangay or '',
            project.project_cost if project.project_cost is not None else '',
            project.source_of_funds or '',
            project.start_date.strftime('%Y-%m-%d') if project.start_date else '',
            project.end_date.strftime('%Y-%m-%d') if project.end_date else '',
            project.get_status_display() or '',
        ])
    
    return response

@login_required
def export_reports_excel(request):
    """Export monitoring reports as Excel"""
    from django.db.models import Q
    
    # Role-based queryset
    if is_head_engineer(request.user) or is_finance_manager(request.user):
        projects = Project.objects.all()
    elif is_project_engineer(request.user):
        projects = Project.objects.filter(assigned_engineers=request.user)
    else:
        projects = Project.objects.none()
    
    # Apply filters from request parameters
    barangay_filter = request.GET.get('barangay', '').strip()
    status_filter = request.GET.get('status', '').strip()
    start_date_filter = request.GET.get('start_date', '').strip()
    end_date_filter = request.GET.get('end_date', '').strip()
    
    # Filter by barangay
    if barangay_filter:
        projects = projects.filter(barangay=barangay_filter)
    
    # Filter by status
    if status_filter:
        # Handle "in_progress" to match both "in_progress" and "ongoing" statuses
        if status_filter == 'in_progress':
            projects = projects.filter(Q(status='in_progress') | Q(status='ongoing'))
        else:
            projects = projects.filter(status=status_filter)
    
    # Filter by start date
    if start_date_filter:
        try:
            from datetime import datetime
            start_date = datetime.strptime(start_date_filter, '%Y-%m-%d').date()
            projects = projects.filter(start_date__gte=start_date)
        except (ValueError, TypeError):
            pass  # Invalid date format, ignore filter
    
    # Filter by end date
    if end_date_filter:
        try:
            from datetime import datetime
            end_date = datetime.strptime(end_date_filter, '%Y-%m-%d').date()
            projects = projects.filter(end_date__lte=end_date)
        except (ValueError, TypeError):
            pass  # Invalid date format, ignore filter
    
    # Filter by source of funds
    source_of_funds_filter = request.GET.get('source_of_funds', '').strip()
    if source_of_funds_filter:
        projects = projects.filter(source_of_funds__iexact=source_of_funds_filter)
    
    # Filter by project cost range
    cost_min_filter = request.GET.get('cost_min', '').strip()
    cost_max_filter = request.GET.get('cost_max', '').strip()
    if cost_min_filter:
        try:
            cost_min = float(cost_min_filter)
            projects = projects.filter(project_cost__gte=cost_min)
        except (ValueError, TypeError):
            pass
    if cost_max_filter:
        try:
            cost_max = float(cost_max_filter)
            projects = projects.filter(project_cost__lte=cost_max)
        except (ValueError, TypeError):
            pass
    
    # Create a new workbook and select the active sheet
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Projects Report"
    
    # Write the header row
    headers = ['#', 'PRN#', 'Name of Project', 'Project Description', 'Barangay', 'Project Cost', 'Source of Funds', 'Date Started', 'Date Ended', 'Status']
    sheet.append(headers)
    
    # Write data rows
    for i, project in enumerate(projects):
        sheet.append([
            i + 1,
            project.prn or '',
            project.name or '',
            project.description or '',
            project.barangay or '',
            project.project_cost if project.project_cost is not None else '',
            project.source_of_funds or '',
            project.start_date.strftime('%Y-%m-%d') if project.start_date else '',
            project.end_date.strftime('%Y-%m-%d') if project.end_date else '',
            project.get_status_display() or '',
        ])
    
    # Create an in-memory BytesIO stream
    excel_file = io.BytesIO()
    workbook.save(excel_file)
    excel_file.seek(0)
    
    # Create the HttpResponse object with the appropriate Excel header
    response = HttpResponse(excel_file.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="projects_report_{timezone.now().strftime("%Y%m%d")}.xlsx"'
    
    return response

@login_required
def export_reports_pdf(request):
    """Export monitoring reports as PDF"""
    from collections import defaultdict
    from django.db.models import Q
    
    # Role-based queryset
    if is_head_engineer(request.user) or is_finance_manager(request.user):
        projects = Project.objects.all()
    elif is_project_engineer(request.user):
        projects = Project.objects.filter(assigned_engineers=request.user)
    else:
        projects = Project.objects.none()
    
    # Apply filters from request parameters
    barangay_filter = request.GET.get('barangay', '').strip()
    status_filter = request.GET.get('status', '').strip()
    start_date_filter = request.GET.get('start_date', '').strip()
    end_date_filter = request.GET.get('end_date', '').strip()
    
    # Filter by barangay
    if barangay_filter:
        projects = projects.filter(barangay=barangay_filter)
    
    # Filter by status
    if status_filter:
        # Handle "in_progress" to match both "in_progress" and "ongoing" statuses
        if status_filter == 'in_progress':
            projects = projects.filter(Q(status='in_progress') | Q(status='ongoing'))
        else:
            projects = projects.filter(status=status_filter)
    
    # Filter by start date
    if start_date_filter:
        try:
            from datetime import datetime
            start_date = datetime.strptime(start_date_filter, '%Y-%m-%d').date()
            projects = projects.filter(start_date__gte=start_date)
        except (ValueError, TypeError):
            pass  # Invalid date format, ignore filter
    
    # Filter by end date
    if end_date_filter:
        try:
            from datetime import datetime
            end_date = datetime.strptime(end_date_filter, '%Y-%m-%d').date()
            projects = projects.filter(end_date__lte=end_date)
        except (ValueError, TypeError):
            pass  # Invalid date format, ignore filter
    
    # Calculate summary statistics
    total_projects = projects.count()
    status_map = defaultdict(int)
    barangay_map = defaultdict(int)
    total_budget = 0
    
    for p in projects:
        # Status counts
        status = p.status
        if status in ['in_progress', 'ongoing']:
            status_map['Ongoing'] += 1
        elif status in ['planned', 'pending']:
            status_map['Planned'] += 1
        elif status == 'completed':
            status_map['Completed'] += 1
        elif status == 'delayed':
            status_map['Delayed'] += 1
        else:
            status_map[status.title()] += 1
        
        # Barangay counts
        if p.barangay:
            barangay_map[p.barangay] += 1
        
        # Budget totals
        if p.project_cost:
            total_budget += float(p.project_cost)
    
    completed_count = status_map.get('Completed', 0)
    ongoing_count = status_map.get('Ongoing', 0)
    planned_count = status_map.get('Planned', 0)
    delayed_count = status_map.get('Delayed', 0)
    
    # Get top barangay
    top_barangay = None
    top_barangay_count = 0
    if barangay_map:
        top_barangay = max(barangay_map.items(), key=lambda x: x[1])
        top_barangay_name = top_barangay[0]
        top_barangay_count = top_barangay[1]
    else:
        top_barangay_name = "N/A"
    
    # If xhtml2pdf is unavailable, return a friendly message
    if pisa is None:
        return HttpResponse('PDF export is temporarily unavailable (missing xhtml2pdf/reportlab).', content_type='text/plain')
    
    # Render the HTML template for the PDF
    template_path = 'monitoring/reports_pdf.html'
    template = get_template(template_path)
    context = {
        'projects': projects,
        'total_projects': total_projects,
        'completed_count': completed_count,
        'ongoing_count': ongoing_count,
        'planned_count': planned_count,
        'delayed_count': delayed_count,
        'total_budget': total_budget,
        'status_map': dict(status_map),
        'barangay_map': dict(barangay_map),
        'top_barangay_name': top_barangay_name,
        'top_barangay_count': top_barangay_count,
        'generated_by': request.user.get_full_name() or request.user.username,
        'generated_at': timezone.now(),
    }
    html = template.render(context)
    
    # Create a PDF
    response = HttpResponse(content_type='application/pdf')
    # Use 'inline' to display in browser, 'attachment' to force download
    response['Content-Disposition'] = f'inline; filename="projects_report_{timezone.now().strftime("%Y%m%d")}.pdf"'
    
    # Create a file-like object to write the PDF data to
    buffer = io.BytesIO()
    
    # Create the PDF object, and write the HTML to it
    pisa_status = pisa.CreatePDF(html, dest=buffer)
    
    # If there were no errors, return the PDF file
    if not pisa_status.err:
        return HttpResponse(buffer.getvalue(), content_type='application/pdf')
    
    # If there were errors, return an error message
    return HttpResponse('We had some errors generating the PDF.', content_type='text/plain')

@login_required
def export_budget_reports_csv(request):
    """Export budget reports as CSV"""
    from projeng.models import ProjectCost
    from collections import defaultdict
    from django.db.models import Q
    
    # Role-based queryset
    if is_head_engineer(request.user) or is_finance_manager(request.user):
        projects = Project.objects.all()
    elif is_project_engineer(request.user):
        projects = Project.objects.filter(assigned_engineers=request.user)
    else:
        projects = Project.objects.none()
    
    # Apply filters (same as budget_reports view)
    selected_barangay = request.GET.get('barangay', '').strip()
    selected_status = request.GET.get('status', '').strip()
    selected_start_date = request.GET.get('start_date', '').strip()
    selected_end_date = request.GET.get('end_date', '').strip()
    selected_cost_min = request.GET.get('cost_min', '').strip()
    selected_cost_max = request.GET.get('cost_max', '').strip()
    selected_budget_status = request.GET.get('budget_status', '').strip()
    
    if selected_barangay:
        projects = projects.filter(barangay=selected_barangay)
    
    if selected_status:
        if selected_status == 'in_progress':
            projects = projects.filter(Q(status='in_progress') | Q(status='ongoing'))
        else:
            projects = projects.filter(status=selected_status)
    
    if selected_start_date:
        projects = projects.filter(start_date__gte=selected_start_date)
    
    if selected_end_date:
        projects = projects.filter(end_date__lte=selected_end_date)

    if selected_cost_min:
        try:
            cost_min_val = float(selected_cost_min)
            projects = projects.filter(project_cost__gte=cost_min_val)
        except (ValueError, TypeError):
            pass
    if selected_cost_max:
        try:
            cost_max_val = float(selected_cost_max)
            projects = projects.filter(project_cost__lte=cost_max_val)
        except (ValueError, TypeError):
            pass
    
    # Create the HttpResponse object with the appropriate CSV header
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = f'attachment; filename="budget_report_{timezone.now().strftime("%Y%m%d")}.csv"'
    
    writer = csv.writer(response)
    # Write the header row
    writer.writerow(['#', 'PRN#', 'Project Name', 'Barangay', 'Budget', 'Spent', 'Remaining', 'Utilization %', 'Status', 'Over/Under Budget'])
    
    # Write data rows
    for i, project in enumerate(projects):
        costs = ProjectCost.objects.filter(project=project)
        spent = sum([float(c.amount) for c in costs]) if costs else 0
        budget = float(project.project_cost) if project.project_cost else 0
        remaining = budget - spent
        utilization = (spent / budget * 100) if budget > 0 else 0
        if budget > 0:
            if spent > budget:
                over_under = 'Over'
            elif utilization >= 80:
                over_under = 'Within'
            else:
                over_under = 'Under'
        else:
            over_under = 'Under'

        if selected_budget_status and over_under.lower() != selected_budget_status.lower():
            continue
        
        writer.writerow([
            i + 1,
            project.prn or '',
            project.name or '',
            project.barangay or '',
            budget,
            spent,
            remaining,
            f'{utilization:.2f}%',
            project.get_status_display() or '',
            over_under,
        ])
    
    return response

@login_required
def export_budget_reports_excel(request):
    """Export budget reports as Excel"""
    from projeng.models import ProjectCost
    from django.db.models import Q
    
    # Role-based queryset
    if is_head_engineer(request.user) or is_finance_manager(request.user):
        projects = Project.objects.all()
    elif is_project_engineer(request.user):
        projects = Project.objects.filter(assigned_engineers=request.user)
    else:
        projects = Project.objects.none()
    
    # Apply filters (same as budget_reports view)
    selected_barangay = request.GET.get('barangay', '').strip()
    selected_status = request.GET.get('status', '').strip()
    selected_start_date = request.GET.get('start_date', '').strip()
    selected_end_date = request.GET.get('end_date', '').strip()
    selected_cost_min = request.GET.get('cost_min', '').strip()
    selected_cost_max = request.GET.get('cost_max', '').strip()
    selected_budget_status = request.GET.get('budget_status', '').strip()
    
    if selected_barangay:
        projects = projects.filter(barangay=selected_barangay)
    
    if selected_status:
        if selected_status == 'in_progress':
            projects = projects.filter(Q(status='in_progress') | Q(status='ongoing'))
        else:
            projects = projects.filter(status=selected_status)
    
    if selected_start_date:
        projects = projects.filter(start_date__gte=selected_start_date)
    
    if selected_end_date:
        projects = projects.filter(end_date__lte=selected_end_date)

    if selected_cost_min:
        try:
            cost_min_val = float(selected_cost_min)
            projects = projects.filter(project_cost__gte=cost_min_val)
        except (ValueError, TypeError):
            pass
    if selected_cost_max:
        try:
            cost_max_val = float(selected_cost_max)
            projects = projects.filter(project_cost__lte=cost_max_val)
        except (ValueError, TypeError):
            pass
    
    # Create a new workbook and select the active sheet
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Budget Report"
    
    # Write the header row
    headers = ['#', 'PRN#', 'Project Name', 'Barangay', 'Budget', 'Spent', 'Remaining', 'Utilization %', 'Status', 'Over/Under Budget']
    sheet.append(headers)
    
    # Write data rows
    for i, project in enumerate(projects):
        costs = ProjectCost.objects.filter(project=project)
        spent = sum([float(c.amount) for c in costs]) if costs else 0
        budget = float(project.project_cost) if project.project_cost else 0
        remaining = budget - spent
        utilization = (spent / budget * 100) if budget > 0 else 0
        if budget > 0:
            if spent > budget:
                over_under = 'Over'
            elif utilization >= 80:
                over_under = 'Within'
            else:
                over_under = 'Under'
        else:
            over_under = 'Under'

        if selected_budget_status and over_under.lower() != selected_budget_status.lower():
            continue
        
        sheet.append([
            i + 1,
            project.prn or '',
            project.name or '',
            project.barangay or '',
            budget,
            spent,
            remaining,
            f'{utilization:.2f}%',
            project.get_status_display() or '',
            over_under,
        ])
    
    # Create an in-memory BytesIO stream
    excel_file = io.BytesIO()
    workbook.save(excel_file)
    excel_file.seek(0)
    
    # Create the HttpResponse object with the appropriate Excel header
    response = HttpResponse(excel_file.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="budget_report_{timezone.now().strftime("%Y%m%d")}.xlsx"'
    
    return response

@login_required
def export_budget_reports_pdf(request):
    """Export budget reports as PDF"""
    from projeng.models import ProjectCost
    from collections import defaultdict
    from django.db.models import Q
    
    # Role-based queryset
    if is_head_engineer(request.user) or is_finance_manager(request.user):
        projects = Project.objects.all()
    elif is_project_engineer(request.user):
        projects = Project.objects.filter(assigned_engineers=request.user)
    else:
        projects = Project.objects.none()
    
    # Apply filters (same as budget_reports view)
    selected_barangay = request.GET.get('barangay', '').strip()
    selected_status = request.GET.get('status', '').strip()
    selected_start_date = request.GET.get('start_date', '').strip()
    selected_end_date = request.GET.get('end_date', '').strip()
    selected_cost_min = request.GET.get('cost_min', '').strip()
    selected_cost_max = request.GET.get('cost_max', '').strip()
    selected_budget_status = request.GET.get('budget_status', '').strip()
    
    if selected_barangay:
        projects = projects.filter(barangay=selected_barangay)
    
    if selected_status:
        if selected_status == 'in_progress':
            projects = projects.filter(Q(status='in_progress') | Q(status='ongoing'))
        else:
            projects = projects.filter(status=selected_status)
    
    if selected_start_date:
        projects = projects.filter(start_date__gte=selected_start_date)
    
    if selected_end_date:
        projects = projects.filter(end_date__lte=selected_end_date)

    if selected_cost_min:
        try:
            cost_min_val = float(selected_cost_min)
            projects = projects.filter(project_cost__gte=cost_min_val)
        except (ValueError, TypeError):
            pass
    if selected_cost_max:
        try:
            cost_max_val = float(selected_cost_max)
            projects = projects.filter(project_cost__lte=cost_max_val)
        except (ValueError, TypeError):
            pass
    
    # Prepare project data (format to match template: #, PRN #, PROJECT, BARANGAY, ASSIGNED ENGINEERS, BUDGET, SPENT, UTILIZATION, OVER/UNDER, STATUS, COST BREAKDOWN)
    project_rows = []
    total_budget = 0.0
    total_spent = 0.0
    over_budget_count = 0
    under_budget_count = 0
    for idx, project in enumerate(projects):
        costs = ProjectCost.objects.filter(project=project)
        spent = sum([float(c.amount) for c in costs]) if costs else 0
        budget = float(project.project_cost) if project.project_cost else 0
        remaining = budget - spent
        utilization = (spent / budget * 100) if budget > 0 else 0
        if budget > 0:
            if spent > budget:
                over_under = 'Over'
            elif utilization >= 80:
                over_under = 'Within'
            else:
                over_under = 'Under'
        else:
            over_under = 'Under'

        # Optional filter by Over/Within/Under
        if selected_budget_status and over_under.lower() != selected_budget_status.lower():
            continue

        total_budget += budget
        total_spent += spent
        if over_under == 'Over':
            over_budget_count += 1
        else:
            under_budget_count += 1

        cost_breakdown_map = defaultdict(float)
        for c in costs:
            cost_breakdown_map[c.get_cost_type_display()] += float(c.amount)
        cost_breakdown_str = '; '.join([f'{k}: ₱{v:,.2f}' for k, v in cost_breakdown_map.items()]) if cost_breakdown_map else '—'
        assigned_engineers = list(project.assigned_engineers.all())
        engineer_names = [eng.get_full_name() or eng.username for eng in assigned_engineers]
        assigned_engineers_str = ', '.join(engineer_names) if engineer_names else '—'

        project_rows.append({
            'Num': idx + 1,
            'PRN': project.prn or '—',
            'Project_Name': project.name or '—',
            'Barangay': project.barangay or '—',
            'Assigned_Engineers': assigned_engineers_str,
            'Budget': f'₱{budget:,.2f}',
            'Spent': f'₱{spent:,.2f}',
            'Utilization': f'{utilization:.2f}%',
            'Over_Under': over_under,
            'Status': project.get_status_display() or '—',
            'Cost_Breakdown': cost_breakdown_str,
        })
    
    # If xhtml2pdf is unavailable, return a friendly message
    if pisa is None:
        return HttpResponse('PDF export is temporarily unavailable (missing xhtml2pdf/reportlab).', content_type='text/plain')
    
    # Render the HTML template for the PDF (use the detailed layout)
    template_path = 'monitoring/budget_report_pdf.html'
    template = get_template(template_path)
    context = {
        'project_data': project_rows,
        'total_projects': len(project_rows),
        'total_budget': total_budget,
        'total_spent': total_spent,
        'over_budget_count': over_budget_count,
        'under_budget_count': under_budget_count,
        'report_date': timezone.now().strftime('%B %d, %Y %H:%M'),
    }
    html = template.render(context)
    
    # Create a PDF
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="budget_report_{timezone.now().strftime("%Y%m%d")}.pdf"'
    
    # Create a file-like object to write the PDF data to
    buffer = io.BytesIO()
    
    # Create the PDF object, and write the HTML to it
    pisa_status = pisa.CreatePDF(html, dest=buffer)
    
    # If there were no errors, return the PDF file
    if not pisa_status.err:
        return HttpResponse(buffer.getvalue(), content_type='application/pdf')
    
    # If there were errors, return an error message
    return HttpResponse('We had some errors generating the PDF.', content_type='text/plain')

@login_required
@head_engineer_required
def head_engineer_notifications(request):
    """View for Head Engineers to manage their notifications"""
    from projeng.models import Notification
    from django.contrib import messages
    from django.shortcuts import redirect
    from django.http import JsonResponse
    from django.core.paginator import Paginator
    from django.db.models import Q
    
    # Get all notifications for the user
    notifications = Notification.objects.filter(recipient=request.user).order_by('-created_at')

    if request.method == 'POST':
        action = request.POST.get('action')
        notification_id = request.POST.get('notification_id')

        if action == 'mark_read' and notification_id:
            try:
                notification = Notification.objects.get(id=notification_id, recipient=request.user)
                notification.is_read = True
                notification.save()
                return JsonResponse({'success': True, 'message': 'Notification marked as read'})
            except Notification.DoesNotExist:
                return JsonResponse({'success': False, 'error': 'Notification not found'})
            except Exception as e:
                return JsonResponse({'success': False, 'error': str(e)})

        elif action == 'mark_all_read':
            try:
                updated_count = notifications.filter(is_read=False).update(is_read=True)
                if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                    return JsonResponse({'success': True, 'message': f'{updated_count} notifications marked as read'})
                messages.success(request, f"All {updated_count} notifications marked as read.")
                return redirect('head_engineer_notifications')
            except Exception as e:
                if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                    return JsonResponse({'success': False, 'error': str(e)})
                messages.error(request, f"Error marking notifications as read: {str(e)}")
                return redirect('head_engineer_notifications')

        elif action == 'delete' and notification_id:
            try:
                notification = Notification.objects.get(id=notification_id, recipient=request.user)
                notification.delete()
                return JsonResponse({'success': True, 'message': 'Notification deleted'})
            except Notification.DoesNotExist:
                return JsonResponse({'success': False, 'error': 'Notification not found'})
            except Exception as e:
                return JsonResponse({'success': False, 'error': str(e)})

        elif action == 'delete_all':
            try:
                deleted_count = notifications.delete()[0]
                if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                    return JsonResponse({'success': True, 'message': f'{deleted_count} notifications deleted'})
                messages.success(request, f"All {deleted_count} notifications deleted.")
                return redirect('head_engineer_notifications')
            except Exception as e:
                if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                    return JsonResponse({'success': False, 'error': str(e)})
                messages.error(request, f"Error deleting notifications: {str(e)}")
                return redirect('head_engineer_notifications')

    # Pagination for large number of notifications
    paginator = Paginator(notifications, 50)  # Show 50 notifications per page
    page_number = request.GET.get('page', 1)
    try:
        page_obj = paginator.get_page(page_number)
    except:
        page_obj = paginator.get_page(1)

    unread_count = notifications.filter(is_read=False).count()
    
    # Add project IDs to notifications for clickable links
    from projeng.utils import get_project_from_notification
    import logging
    logger = logging.getLogger(__name__)
    notifications_with_projects = []
    for notification in page_obj:
        project_id = None
        try:
            if notification.message:
                # Check if it's a Budget Concern notification
                if 'Budget Concern' in notification.message or '📋' in notification.message:
                    logger.info(f"Processing Budget Concern notification {notification.id}: {notification.message[:100]}")
                project_id = get_project_from_notification(notification.message)
                if project_id:
                    logger.info(f"Extracted project_id={project_id} for notification {notification.id}")
                else:
                    if 'Budget Concern' in notification.message or '📋' in notification.message:
                        logger.warning(f"Failed to extract project_id for Budget Concern notification {notification.id}")
        except Exception as e:
            # Log error but don't break the page
            logger.error(f"Error getting project from notification {notification.id}: {str(e)}", exc_info=True)
        notifications_with_projects.append({
            'notification': notification,
            'project_id': project_id
        })

    context = {
        'notifications': page_obj,  # Use paginated notifications
        'notifications_with_projects': notifications_with_projects,  # Notifications with project IDs
        'unread_count': unread_count,
        'page_obj': page_obj,  # For pagination controls
    }
    return render(request, 'monitoring/notifications.html', context)

@login_required
@head_engineer_required
def export_project_comprehensive_pdf(request, pk):
    """Export comprehensive project report (Progress + Budget) as PDF"""
    from projeng.models import Project, ProjectProgress, ProjectCost, BudgetRequest
    from django.utils import timezone
    from datetime import timedelta
    
    try:
        project = Project.objects.get(pk=pk)
    except Project.DoesNotExist:
        return HttpResponse('Project not found.', status=404)
    
    # Get all progress updates (include related photos as means of verification)
    progress_updates = (
        ProjectProgress.objects
        .filter(project=project)
        .prefetch_related('photos')
        .select_related('created_by')
        .order_by('date', 'id')
        .distinct()
    )
    
    # Get all cost entries
    costs = ProjectCost.objects.filter(project=project).order_by('date')
    
    # Calculate analytics
    latest_progress = progress_updates.last() if progress_updates else None
    total_progress = latest_progress.percentage_complete if latest_progress else 0
    total_cost = sum([float(c.amount) for c in costs]) if costs else 0
    budget = float(project.project_cost) if project.project_cost else 0
    remaining_budget = budget - total_cost
    budget_utilization = (total_cost / budget * 100) if budget > 0 else 0

    # Progress vs budget metrics (panel: efficiency/performance ratios)
    budget_used_pct = float(budget_utilization) if budget_utilization else 0.0
    efficiency_ratio = (float(total_progress) / budget_used_pct) if budget_used_pct > 0 else None
    project_budget_ratio = (float(total_cost) / float(budget)) if budget > 0 else None
    
    # Calculate timeline
    today = timezone.now().date()
    days_elapsed = (today - project.start_date).days if project.start_date else 0
    total_days = (project.end_date - project.start_date).days if project.start_date and project.end_date else 0
    days_remaining = total_days - days_elapsed if total_days > 0 else 0

    expected_progress = None
    progress_variance = None
    performance_ratio = None
    if project.start_date and project.end_date and total_days > 0:
        elapsed_days = max(0, min(total_days, days_elapsed))
        expected_progress = min(100.0, (elapsed_days / total_days) * 100.0)
        progress_variance = float(total_progress) - float(expected_progress)
        performance_ratio = (float(total_progress) / float(expected_progress)) if expected_progress > 0 else None

    # Simple status labels for report readability
    budget_status_label = 'UNDER BUDGET'
    if budget_used_pct > 100:
        budget_status_label = 'OVER BUDGET'
    elif budget_used_pct >= 90:
        budget_status_label = 'AT RISK'

    performance_label = 'N/A'
    if performance_ratio is not None:
        if performance_ratio >= 1.05:
            performance_label = 'Ahead of schedule'
        elif performance_ratio >= 0.95:
            performance_label = 'On schedule'
        else:
            performance_label = 'Behind schedule'

    # Panel #29: Achieve Satisfaction (0-100) based on schedule + budget alignment
    satisfaction_score = None
    satisfaction_label = 'N/A'
    try:
        budget_component = None
        if efficiency_ratio is not None:
            budget_component = max(0.0, 100.0 - min(100.0, abs(1.0 - float(efficiency_ratio)) * 100.0))

        schedule_component = None
        if progress_variance is not None:
            schedule_component = max(0.0, 100.0 - abs(float(progress_variance)))

        if budget_component is not None and schedule_component is not None:
            base = (0.5 * schedule_component) + (0.5 * budget_component)
        elif budget_component is not None:
            base = budget_component
        else:
            base = None

        if base is not None and budget_used_pct > 100:
            base = base - 10.0  # over-budget penalty

        if base is not None:
            satisfaction_score = round(max(0.0, min(100.0, float(base))), 1)
            if satisfaction_score >= 85:
                satisfaction_label = 'High'
            elif satisfaction_score >= 70:
                satisfaction_label = 'Moderate'
            else:
                satisfaction_label = 'Low'
    except Exception:
        satisfaction_score = None
        satisfaction_label = 'N/A'
    
    # Cost breakdown by category
    from collections import defaultdict
    cost_breakdown = defaultdict(float)
    for cost in costs:
        cost_breakdown[cost.get_cost_type_display()] += float(cost.amount)
    
    # Assigned engineers
    assigned_engineers = project.assigned_engineers.all()

    # Budget requests (additional budget) history
    budget_requests = (
        BudgetRequest.objects
        .filter(project=project)
        .select_related('requested_by', 'reviewed_by')
        .prefetch_related('attachments')
        .order_by('created_at')
    )

    # Uploaded images for PDF (xhtml2pdf works best with URLs, not data URIs) - 4 per row
    from projeng.models import ProjectDocument
    all_docs = ProjectDocument.objects.filter(project=project).select_related('uploaded_by').order_by('-uploaded_at')
    image_extensions = ['.jpg', '.jpeg', '.png', '.gif', '.webp', '.bmp']
    image_documents = [d for d in all_docs if d.file and any((d.file.name or '').lower().endswith(ext) for ext in image_extensions)]
    flat_images = _build_uploaded_images_for_pdf_urls(request, progress_updates, image_documents, project)
    COLS = 4
    uploaded_images_rows = []
    for i in range(0, len(flat_images), COLS):
        row = flat_images[i:i + COLS]
        row.extend([{'url': '', 'caption': ''}] * (COLS - len(row)))  # pad to 5 columns
        uploaded_images_rows.append(row)
    
    # Render the HTML template for the PDF
    template_path = 'monitoring/project_comprehensive_report_pdf.html'
    template = get_template(template_path)
    context = {
        'project': project,
        'progress_updates': progress_updates,
        'costs': costs,
        'latest_progress': latest_progress,
        'total_progress': total_progress,
        'total_cost': total_cost,
        'budget': budget,
        'remaining_budget': remaining_budget,
        'budget_utilization': budget_utilization,
        'budget_used_pct': budget_used_pct,
        'project_budget_ratio': project_budget_ratio,
        'efficiency_ratio': efficiency_ratio,
        'expected_progress': expected_progress,
        'progress_variance': progress_variance,
        'performance_ratio': performance_ratio,
        'budget_status_label': budget_status_label,
        'performance_label': performance_label,
        'satisfaction_score': satisfaction_score,
        'satisfaction_label': satisfaction_label,
        'days_elapsed': days_elapsed,
        'total_days': total_days,
        'days_remaining': days_remaining,
        'cost_breakdown': dict(cost_breakdown),
        'assigned_engineers': assigned_engineers,
        'budget_requests': budget_requests,
        'has_progress_photos': any(getattr(u, 'photos', None) and u.photos.all() for u in progress_updates),
        'uploaded_images_rows': uploaded_images_rows,
        'generated_by': request.user.get_full_name() or request.user.username,
        'generated_at': timezone.now(),
    }
    html = template.render(context)

    # Use wkhtmltopdf (pdfkit) if available, else fall back to xhtml2pdf
    pdf_bytes = None
    if pdfkit is not None:
        try:
            options = {
                'page-size': 'A4',
                'margin-top': '20mm',
                'margin-right': '20mm',
                'margin-bottom': '24mm',
                'margin-left': '20mm',
                'encoding': 'UTF-8',
                'footer-center': 'Page [page] of [topage]',
                'footer-font-size': '8',
            }
            pdf_bytes = pdfkit.from_string(html, False, options=options)
        except Exception as e:
            import logging
            logging.getLogger(__name__).warning('wkhtmltopdf failed: %s, falling back to xhtml2pdf', str(e))
            pdf_bytes = None

    if pdf_bytes is None and pisa is not None:
        result = io.BytesIO()
        pdf = pisa.CreatePDF(io.BytesIO(html.encode("UTF-8")), result)
        if not pdf.err:
            pdf_bytes = result.getvalue()

    if pdf_bytes:
        response = HttpResponse(pdf_bytes, content_type='application/pdf')
        filename = f"project_report_{project.prn or project.id}_{timezone.now().strftime('%Y%m%d')}.pdf"
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        return response
    if pisa is None and pdfkit is None:
        return HttpResponse('PDF export unavailable. Install wkhtmltopdf and pdfkit, or xhtml2pdf.', content_type='text/plain')
    return HttpResponse('Error generating PDF', content_type='text/plain')

@login_required
@head_engineer_required
def export_project_comprehensive_excel(request, pk):
    """Export comprehensive project report (Progress + Budget) as Excel"""
    from projeng.models import Project, ProjectProgress, ProjectCost, BudgetRequest
    from django.utils import timezone
    from collections import defaultdict
    
    try:
        project = Project.objects.get(pk=pk)
    except Project.DoesNotExist:
        return HttpResponse('Project not found.', status=404)
    
    # Get all progress updates (include related photos as means of verification)
    progress_updates = (
        ProjectProgress.objects
        .filter(project=project)
        .prefetch_related('photos')
        .select_related('created_by')
        .order_by('date', 'id')
        .distinct()
    )
    
    # Get all cost entries
    costs = ProjectCost.objects.filter(project=project).order_by('date')
    
    # Calculate analytics
    latest_progress = progress_updates.last() if progress_updates else None
    total_progress = latest_progress.percentage_complete if latest_progress else 0
    total_cost = sum([float(c.amount) for c in costs]) if costs else 0
    budget = float(project.project_cost) if project.project_cost else 0
    remaining_budget = budget - total_cost
    budget_utilization = (total_cost / budget * 100) if budget > 0 else 0

    # Ratios + satisfaction (same definitions as PDF)
    budget_used_pct = float(budget_utilization) if budget_utilization else 0.0
    efficiency_ratio = (float(total_progress) / budget_used_pct) if budget_used_pct > 0 else None

    today = timezone.now().date()
    days_elapsed = (today - project.start_date).days if project.start_date else 0
    total_days = (project.end_date - project.start_date).days if project.start_date and project.end_date else 0
    expected_progress = None
    progress_variance = None
    performance_ratio = None
    if project.start_date and project.end_date and total_days > 0:
        elapsed_days = max(0, min(total_days, days_elapsed))
        expected_progress = min(100.0, (elapsed_days / total_days) * 100.0)
        progress_variance = float(total_progress) - float(expected_progress)
        performance_ratio = (float(total_progress) / float(expected_progress)) if expected_progress > 0 else None

    satisfaction_score = None
    satisfaction_label = 'N/A'
    try:
        budget_component = None
        if efficiency_ratio is not None:
            budget_component = max(0.0, 100.0 - min(100.0, abs(1.0 - float(efficiency_ratio)) * 100.0))
        schedule_component = None
        if progress_variance is not None:
            schedule_component = max(0.0, 100.0 - abs(float(progress_variance)))
        if budget_component is not None and schedule_component is not None:
            base = (0.5 * schedule_component) + (0.5 * budget_component)
        elif budget_component is not None:
            base = budget_component
        else:
            base = None
        if base is not None and budget_used_pct > 100:
            base = base - 10.0
        if base is not None:
            satisfaction_score = round(max(0.0, min(100.0, float(base))), 1)
            if satisfaction_score >= 85:
                satisfaction_label = 'High'
            elif satisfaction_score >= 70:
                satisfaction_label = 'Moderate'
            else:
                satisfaction_label = 'Low'
    except Exception:
        satisfaction_score = None
        satisfaction_label = 'N/A'
    
    # Create Excel workbook
    wb = openpyxl.Workbook()
    
    # Sheet 1: Project Overview
    ws1 = wb.active
    ws1.title = "Project Overview"
    ws1.append(['PROJECT COMPREHENSIVE REPORT'])
    ws1.append(['Project Name', project.name])
    ws1.append(['PRN Number', project.prn or 'N/A'])
    ws1.append(['Location', project.barangay or 'N/A'])
    ws1.append(['Status', project.get_status_display()])
    ws1.append(['Start Date', str(project.start_date) if project.start_date else 'N/A'])
    ws1.append(['End Date', str(project.end_date) if project.end_date else 'N/A'])
    ws1.append(['Assigned Engineers', ', '.join([e.get_full_name() or e.username for e in project.assigned_engineers.all()])])
    ws1.append([])
    ws1.append(['PROGRESS SUMMARY'])
    ws1.append(['Overall Progress', f'{total_progress}%'])
    ws1.append(['BUDGET SUMMARY'])
    ws1.append(['Total Budget', f'₱{budget:,.2f}'])
    ws1.append(['Total Spent', f'₱{total_cost:,.2f}'])
    ws1.append(['Remaining', f'₱{remaining_budget:,.2f}'])
    ws1.append(['Utilization', f'{budget_utilization:.2f}%'])
    ws1.append([])
    ws1.append(['PERFORMANCE & SATISFACTION'])
    ws1.append(['Expected Progress (Timeline)', f'{expected_progress:.2f}%' if expected_progress is not None else 'N/A'])
    ws1.append(['Progress Variance', f'{progress_variance:.2f}%' if progress_variance is not None else 'N/A'])
    ws1.append(['Performance Ratio (Actual ÷ Expected)', f'{performance_ratio:.2f}' if performance_ratio is not None else 'N/A'])
    ws1.append(['Efficiency Ratio (Progress ÷ Budget Used %)', f'{float(efficiency_ratio):.2f}' if efficiency_ratio is not None else 'N/A'])
    ws1.append(['Achieve Satisfaction', f'{satisfaction_score:.1f} ({satisfaction_label})' if satisfaction_score is not None else 'N/A'])
    
    # Sheet 2: Progress Details
    ws2 = wb.create_sheet("Progress Details")
    ws2.append(['Date', 'Progress %', 'Description', 'Engineer'])
    for update in progress_updates:
        engineer_name = update.created_by.get_full_name() or update.created_by.username if update.created_by else 'N/A'
        ws2.append([
            str(update.date),
            update.percentage_complete,
            update.description or '',
            engineer_name
        ])

    # Sheet 2b: Progress Photos (Means of Verification)
    ws2b = wb.create_sheet("Progress Photos")
    ws2b.append(['Update Date', 'Progress %', 'Photo Uploaded At', 'Photo File', 'Photo URL'])
    for update in progress_updates:
        photos_qs = getattr(update, 'photos', None)
        photos = list(photos_qs.all()) if photos_qs is not None else []
        if not photos:
            continue
        for photo in photos:
            try:
                url = photo.image.url
            except Exception:
                url = ''
            ws2b.append([
                str(update.date),
                update.percentage_complete,
                timezone.localtime(photo.uploaded_at).strftime('%Y-%m-%d %H:%M') if photo.uploaded_at else '',
                getattr(photo.image, 'name', ''),
                url,
            ])
    
    # Sheet 3: Budget Details
    ws3 = wb.create_sheet("Budget Details")
    ws3.append(['Date', 'Category', 'Description', 'Amount'])
    for cost in costs:
        ws3.append([
            str(cost.date),
            cost.get_cost_type_display(),
            cost.description or '',
            float(cost.amount)
        ])
    ws3.append([])
    ws3.append(['TOTAL', '', '', f'₱{total_cost:,.2f}'])
    
    # Sheet 4: Cost Breakdown
    ws4 = wb.create_sheet("Cost Breakdown")
    cost_breakdown = defaultdict(float)
    for cost in costs:
        cost_breakdown[cost.get_cost_type_display()] += float(cost.amount)
    ws4.append(['Category', 'Amount'])
    for category, amount in cost_breakdown.items():
        ws4.append([category, f'₱{amount:,.2f}'])
    ws4.append(['TOTAL', f'₱{total_cost:,.2f}'])
    
    # Sheet 5: Progress vs Budget
    ws5 = wb.create_sheet("Progress vs Budget")
    ws5.append(['Date', 'Progress %', 'Budget Used %', 'Efficiency Ratio'])
    for update in progress_updates:
        # Calculate budget used up to this date
        costs_up_to_date = ProjectCost.objects.filter(project=project, date__lte=update.date)
        budget_used = sum([float(c.amount) for c in costs_up_to_date]) if costs_up_to_date else 0
        budget_used_pct = (budget_used / budget * 100) if budget > 0 else 0
        efficiency = (update.percentage_complete / budget_used_pct) if budget_used_pct > 0 else 0
        ws5.append([
            str(update.date),
            update.percentage_complete,
            f'{budget_used_pct:.2f}%',
            f'{efficiency:.2f}'
        ])

    # Sheet 6: Budget Requests
    ws6 = wb.create_sheet("Budget Requests")
    ws6.append(['Created At', 'Requested Amount', 'Status', 'Approved Amount', 'Requested By', 'Reviewed By', 'Reviewed At', 'Decision Notes', 'Proof Count'])
    budget_requests = (
        BudgetRequest.objects
        .filter(project=project)
        .select_related('requested_by', 'reviewed_by')
        .prefetch_related('attachments')
        .order_by('created_at')
    )
    for br in budget_requests:
        ws6.append([
            timezone.localtime(br.created_at).strftime('%Y-%m-%d %H:%M') if br.created_at else '',
            float(br.requested_amount) if br.requested_amount else 0,
            br.get_status_display(),
            float(br.approved_amount) if br.approved_amount else '',
            (br.requested_by.get_full_name() or br.requested_by.username) if br.requested_by else '',
            (br.reviewed_by.get_full_name() or br.reviewed_by.username) if br.reviewed_by else '',
            timezone.localtime(br.reviewed_at).strftime('%Y-%m-%d %H:%M') if br.reviewed_at else '',
            br.decision_notes or '',
            br.attachments.count() if hasattr(br, 'attachments') else 0,
        ])
    
    # Save to response
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    filename = f"project_report_{project.prn or project.id}_{timezone.now().strftime('%Y%m%d')}.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response

@login_required
@head_engineer_required
def export_project_comprehensive_csv(request, pk):
    """Export comprehensive project report (Progress + Budget) as CSV"""
    from projeng.models import Project, ProjectProgress, ProjectCost
    from django.utils import timezone
    
    try:
        project = Project.objects.get(pk=pk)
    except Project.DoesNotExist:
        return HttpResponse('Project not found.', status=404)
    
    # Get all progress updates (include related photos as means of verification)
    progress_updates = (
        ProjectProgress.objects
        .filter(project=project)
        .prefetch_related('photos')
        .select_related('created_by')
        .order_by('date', 'id')
        .distinct()
    )
    
    # Get all cost entries
    costs = ProjectCost.objects.filter(project=project).order_by('date')
    
    # Calculate analytics
    latest_progress = progress_updates.last() if progress_updates else None
    total_progress = latest_progress.percentage_complete if latest_progress else 0
    total_cost = sum([float(c.amount) for c in costs]) if costs else 0
    budget = float(project.project_cost) if project.project_cost else 0
    remaining_budget = budget - total_cost
    budget_utilization = (total_cost / budget * 100) if budget > 0 else 0

    # Ratios + satisfaction (same definitions as PDF)
    budget_used_pct = float(budget_utilization) if budget_utilization else 0.0
    efficiency_ratio = (float(total_progress) / budget_used_pct) if budget_used_pct > 0 else None

    today = timezone.now().date()
    days_elapsed = (today - project.start_date).days if project.start_date else 0
    total_days = (project.end_date - project.start_date).days if project.start_date and project.end_date else 0
    expected_progress = None
    progress_variance = None
    performance_ratio = None
    if project.start_date and project.end_date and total_days > 0:
        elapsed_days = max(0, min(total_days, days_elapsed))
        expected_progress = min(100.0, (elapsed_days / total_days) * 100.0)
        progress_variance = float(total_progress) - float(expected_progress)
        performance_ratio = (float(total_progress) / float(expected_progress)) if expected_progress > 0 else None

    satisfaction_score = None
    satisfaction_label = 'N/A'
    try:
        budget_component = None
        if efficiency_ratio is not None:
            budget_component = max(0.0, 100.0 - min(100.0, abs(1.0 - float(efficiency_ratio)) * 100.0))
        schedule_component = None
        if progress_variance is not None:
            schedule_component = max(0.0, 100.0 - abs(float(progress_variance)))
        if budget_component is not None and schedule_component is not None:
            base = (0.5 * schedule_component) + (0.5 * budget_component)
        elif budget_component is not None:
            base = budget_component
        else:
            base = None
        if base is not None and budget_used_pct > 100:
            base = base - 10.0
        if base is not None:
            satisfaction_score = round(max(0.0, min(100.0, float(base))), 1)
            if satisfaction_score >= 85:
                satisfaction_label = 'High'
            elif satisfaction_score >= 70:
                satisfaction_label = 'Moderate'
            else:
                satisfaction_label = 'Low'
    except Exception:
        satisfaction_score = None
        satisfaction_label = 'N/A'
    
    # Create CSV response
    response = HttpResponse(content_type='text/csv')
    filename = f"project_report_{project.prn or project.id}_{timezone.now().strftime('%Y%m%d')}.csv"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    writer = csv.writer(response)
    
    # Project Information
    writer.writerow(['PROJECT COMPREHENSIVE REPORT'])
    writer.writerow(['Project Name', project.name])
    writer.writerow(['PRN Number', project.prn or 'N/A'])
    writer.writerow(['Location', project.barangay or 'N/A'])
    writer.writerow(['Status', project.get_status_display()])
    writer.writerow(['Start Date', str(project.start_date) if project.start_date else 'N/A'])
    writer.writerow(['End Date', str(project.end_date) if project.end_date else 'N/A'])
    writer.writerow([])
    
    # Progress Summary
    writer.writerow(['PROGRESS SUMMARY'])
    writer.writerow(['Overall Progress', f'{total_progress}%'])
    writer.writerow([])
    
    # Budget Summary
    writer.writerow(['BUDGET SUMMARY'])
    writer.writerow(['Total Budget', f'₱{budget:,.2f}'])
    writer.writerow(['Total Spent', f'₱{total_cost:,.2f}'])
    writer.writerow(['Remaining', f'₱{remaining_budget:,.2f}'])
    writer.writerow(['Utilization', f'{budget_utilization:.2f}%'])
    writer.writerow([])
    writer.writerow(['PERFORMANCE & SATISFACTION'])
    writer.writerow(['Expected Progress (Timeline)', f'{expected_progress:.2f}%' if expected_progress is not None else 'N/A'])
    writer.writerow(['Progress Variance', f'{progress_variance:.2f}%' if progress_variance is not None else 'N/A'])
    writer.writerow(['Performance Ratio (Actual ÷ Expected)', f'{performance_ratio:.2f}' if performance_ratio is not None else 'N/A'])
    writer.writerow(['Efficiency Ratio (Progress ÷ Budget Used %)', f'{float(efficiency_ratio):.2f}' if efficiency_ratio is not None else 'N/A'])
    writer.writerow(['Achieve Satisfaction', f'{satisfaction_score:.1f} ({satisfaction_label})' if satisfaction_score is not None else 'N/A'])
    writer.writerow([])
    
    # Progress Details
    writer.writerow(['PROGRESS DETAILS'])
    writer.writerow(['Date', 'Progress %', 'Description', 'Engineer', 'Photo Count', 'Photo Files'])
    for update in progress_updates:
        engineer_name = update.created_by.get_full_name() or update.created_by.username if update.created_by else 'N/A'
        photos_qs = getattr(update, 'photos', None)
        photos = list(photos_qs.all()) if photos_qs is not None else []
        photo_files = ', '.join([getattr(p.image, 'name', '') for p in photos if getattr(p, 'image', None)])
        writer.writerow([
            str(update.date),
            update.percentage_complete,
            update.description or '',
            engineer_name,
            len(photos),
            photo_files,
        ])
    writer.writerow([])
    
    # Budget Details
    writer.writerow(['BUDGET DETAILS'])
    writer.writerow(['Date', 'Category', 'Description', 'Amount'])
    for cost in costs:
        writer.writerow([
            str(cost.date),
            cost.get_cost_type_display(),
            cost.description or '',
            float(cost.amount)
        ])
    writer.writerow([])
    writer.writerow(['TOTAL', '', '', f'₱{total_cost:,.2f}'])
    
    return response
